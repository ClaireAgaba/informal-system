from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import AssessmentSeries
from .serializers import AssessmentSeriesSerializer


class AssessmentSeriesViewSet(viewsets.ModelViewSet):
    """ViewSet for AssessmentSeries model"""
    queryset = AssessmentSeries.objects.all()
    serializer_class = AssessmentSeriesSerializer
    permission_classes = [permissions.AllowAny]  # Temporarily allow unauthenticated access
    
    @action(detail=True, methods=['post'])
    def set_current(self, request, pk=None):
        """Set this series as the current assessment series"""
        series = self.get_object()
        series.is_current = True
        series.save()
        serializer = self.get_serializer(series)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def release_results(self, request, pk=None):
        """Release results for this assessment series"""
        series = self.get_object()
        series.results_released = True
        series.save()
        serializer = self.get_serializer(series)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def export_special_needs(self, request, pk=None):
        """Export special needs and refugee candidates for a specific series into an Excel file"""
        from django.http import HttpResponse
        from datetime import date
        from django.db.models import Q
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from candidates.models import CandidateEnrollment

        series = self.get_object()

        # Find candidates enrolled in this series who are either refugees or have a disability
        # Use select_related to optimize the queries
        enrollments = CandidateEnrollment.objects.filter(
            assessment_series=series,
            is_active=True
        ).filter(
            Q(candidate__has_disability=True) | Q(candidate__is_refugee=True)
        ).select_related(
            'candidate',
            'candidate__assessment_center',
            'candidate__occupation',
            'candidate__occupation__sector',
            'candidate__district',
            'candidate__nature_of_disability',
            'occupation_level'
        )

        wb = openpyxl.Workbook()
        
        # --- Sheet 1: Special Needs Candidates ---
        ws_special = wb.active
        ws_special.title = "Special Needs Candidates"
        
        special_headers = [
            ('Reg No', 20), ('Full Name', 25), ('Center No', 15), ('Center Name', 30), ('Center Contact', 15),
            ('Category', 12), ('Occupation', 20), ('Sector', 15), ('Level/Modules', 20),
            ('Nature of Disability', 25), ('Disability Notes', 30),
            ('Age', 6), ('District', 15), ('Gender', 8), ('Contact', 15)
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for col, (header, width) in enumerate(special_headers, 1):
            cell = ws_special.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws_special.column_dimensions[cell.column_letter].width = width

        # --- Sheet 2: Refugees ---
        ws_refugees = wb.create_sheet(title="Refugees")
        
        refugee_headers = [
            ('Reg No', 20), ('Full Name', 25), ('Center No', 15), ('Center Name', 30), ('Center Contact', 15),
            ('Category', 12), ('Occupation', 20), ('Sector', 15), ('Level/Modules', 20),
            ('Refugee ID', 20), ('Nationality', 15),
            ('Age', 6), ('District', 15), ('Gender', 8), ('Contact', 15)
        ]
        
        for col, (header, width) in enumerate(refugee_headers, 1):
            cell = ws_refugees.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws_refugees.column_dimensions[cell.column_letter].width = width
            
        category_map = {'modular': 'Modular', 'formal': 'Formal', 'workers_pas': "Worker's PAS"}
        gender_map = {'male': 'Male', 'female': 'Female', 'other': 'Other'}
        
        def calc_age(dob):
            if not dob:
                return ''
            today = date.today()
            return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

        special_row = 2
        refugee_row = 2

        # Filter enrollments into the two lists
        # A candidate might be both and will appear on both sheets
        for enrollment in enrollments:
            c = enrollment.candidate
            
            # Formatting level/modules
            level_modules = ''
            if c.registration_category in ['formal', 'workers_pas'] and enrollment.occupation_level:
                level_modules = enrollment.occupation_level.level_name
            elif c.registration_category in ['modular', 'workers_pas']:
                modules = enrollment.modules.all()
                if modules:
                    level_modules = ', '.join([m.module.module_code for m in modules])
                elif c.registration_category == 'workers_pas':
                    papers = enrollment.papers.all()
                    if papers:
                        level_modules = ', '.join([p.paper.paper_code for p in papers])

            if c.has_disability:
                ws_special.cell(row=special_row, column=1, value=c.registration_number or '')
                ws_special.cell(row=special_row, column=2, value=c.full_name or '')
                ws_special.cell(row=special_row, column=3, value=c.assessment_center.center_number if c.assessment_center else '')
                ws_special.cell(row=special_row, column=4, value=c.assessment_center.center_name if c.assessment_center else '')
                ws_special.cell(row=special_row, column=5, value=c.assessment_center.contact_1 if c.assessment_center else '')
                ws_special.cell(row=special_row, column=6, value=category_map.get(c.registration_category, ''))
                ws_special.cell(row=special_row, column=7, value=c.occupation.occ_name if c.occupation else '')
                ws_special.cell(row=special_row, column=8, value=c.occupation.sector.name if c.occupation and c.occupation.sector else '')
                ws_special.cell(row=special_row, column=9, value=level_modules)
                ws_special.cell(row=special_row, column=10, value=c.nature_of_disability.name if c.nature_of_disability else '')
                ws_special.cell(row=special_row, column=11, value=c.disability_specification or '')
                ws_special.cell(row=special_row, column=12, value=calc_age(c.date_of_birth))
                ws_special.cell(row=special_row, column=13, value=c.district.name if c.district else '')
                ws_special.cell(row=special_row, column=14, value=gender_map.get(c.gender, ''))
                ws_special.cell(row=special_row, column=15, value=c.contact or '')
                special_row += 1
                
            if c.is_refugee:
                ws_refugees.cell(row=refugee_row, column=1, value=c.registration_number or '')
                ws_refugees.cell(row=refugee_row, column=2, value=c.full_name or '')
                ws_refugees.cell(row=refugee_row, column=3, value=c.assessment_center.center_number if c.assessment_center else '')
                ws_refugees.cell(row=refugee_row, column=4, value=c.assessment_center.center_name if c.assessment_center else '')
                ws_refugees.cell(row=refugee_row, column=5, value=c.assessment_center.contact_1 if c.assessment_center else '')
                ws_refugees.cell(row=refugee_row, column=6, value=category_map.get(c.registration_category, ''))
                ws_refugees.cell(row=refugee_row, column=7, value=c.occupation.occ_name if c.occupation else '')
                ws_refugees.cell(row=refugee_row, column=8, value=c.occupation.sector.name if c.occupation and c.occupation.sector else '')
                ws_refugees.cell(row=refugee_row, column=9, value=level_modules)
                ws_refugees.cell(row=refugee_row, column=10, value=c.refugee_number or '')
                ws_refugees.cell(row=refugee_row, column=11, value=c.nationality or 'Uganda')
                ws_refugees.cell(row=refugee_row, column=12, value=calc_age(c.date_of_birth))
                ws_refugees.cell(row=refugee_row, column=13, value=c.district.name if c.district else '')
                ws_refugees.cell(row=refugee_row, column=14, value=gender_map.get(c.gender, ''))
                ws_refugees.cell(row=refugee_row, column=15, value=c.contact or '')
                refugee_row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_series_name = str(series.name).replace(" ", "_").lower()
        response['Content-Disposition'] = f'attachment; filename=special_needs_refugees_{safe_series_name}_{date.today().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        
        return response

    @action(detail=True, methods=['get'])
    def export_registration_summary(self, request, pk=None):
        """
        Export registration summary grouped by center and occupation.
        Creates multiple sheets:
        - Level 1, Level 2, Level 3, Level 4 (for Formal candidates)
        - Modular
        - Workers PAS
        """
        from django.http import HttpResponse
        from datetime import date
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from candidates.models import CandidateEnrollment
        from collections import defaultdict
        from decimal import Decimal

        series = self.get_object()

        # Fetch active enrollments with related fields
        enrollments = CandidateEnrollment.objects.filter(
            assessment_series=series,
            is_active=True
        ).select_related(
            'candidate',
            'candidate__occupation',
            'candidate__assessment_center',
            'occupation_level'
        ).prefetch_related(
            'modules__module',
            'papers__paper'
        )

        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define sheet configurations
        # Formal levels - will match level names that START with these prefixes
        formal_level_prefixes = ['Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5']
        
        # Headers for formal sheets (with level column)
        formal_headers = [
            ('Center Code', 15),
            ('Center Name', 40),
            ('Occ Code', 12),
            ('Occ Name', 30),
            ('Reg Category', 15),
            ('Level', 12),
            ('No of Candidates', 18),
            ('Amount Billed', 18),
        ]
        
        # Headers for modular sheets (aggregate by center/occ, no module breakdown to avoid duplication)
        modular_headers = [
            ('Center Code', 15),
            ('Center Name', 40),
            ('Occ Code', 12),
            ('Occ Name', 30),
            ('Reg Category', 15),
            ('No of Candidates', 18),
            ('Amount Billed', 18),
        ]
        
        # Headers for workers_pas sheets
        workers_pas_headers = [
            ('Center Code', 15),
            ('Center Name', 40),
            ('Occ Code', 12),
            ('Occ Name', 30),
            ('Reg Category', 15),
            ('No of Candidates', 18),
            ('Amount Billed', 18),
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        def create_sheet_with_headers(wb, sheet_name, headers):
            ws = wb.create_sheet(title=sheet_name)
            for col, (header, width) in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                ws.column_dimensions[cell.column_letter].width = width
            return ws
        
        # Data structures to aggregate by center/occupation
        # For formal: key = (center_code, center_name, occ_code, occ_name, level)
        # For modular: key = (center_code, center_name, occ_code, occ_name) - no module breakdown to avoid duplication
        # For workers_pas: key = (center_code, center_name, occ_code, occ_name)
        formal_data = defaultdict(lambda: {'count': 0, 'amount': Decimal('0.00')})
        modular_data = defaultdict(lambda: {'count': 0, 'amount': Decimal('0.00')})
        workers_pas_data = defaultdict(lambda: {'count': 0, 'amount': Decimal('0.00')})
        
        for enrollment in enrollments:
            c = enrollment.candidate
            center = c.assessment_center
            center_code = center.center_number if center else 'N/A'
            center_name = center.center_name if center else 'N/A'
            occ_code = c.occupation.occ_code if c.occupation else 'N/A'
            occ_name = c.occupation.occ_name if c.occupation else 'N/A'
            amount = enrollment.total_amount or Decimal('0.00')
            
            if c.is_formal():
                level_name = enrollment.occupation_level.level_name if enrollment.occupation_level else 'N/A'
                key = (center_code, center_name, occ_code, occ_name, level_name)
                formal_data[key]['count'] += 1
                formal_data[key]['amount'] += amount
            elif c.is_modular():
                # Count each enrollment once (not per module) to avoid duplication
                key = (center_code, center_name, occ_code, occ_name)
                modular_data[key]['count'] += 1
                modular_data[key]['amount'] += amount
            elif c.is_workers_pas():
                key = (center_code, center_name, occ_code, occ_name)
                workers_pas_data[key]['count'] += 1
                workers_pas_data[key]['amount'] += amount
        
        # Helper function to get level prefix (e.g., "Level 1 CK" -> "Level 1")
        def get_level_prefix(level_name):
            for prefix in formal_level_prefixes:
                if level_name and level_name.startswith(prefix):
                    return prefix
            return level_name  # Return as-is if no match
        
        # Group formal data by level prefix
        formal_by_level = defaultdict(list)
        for (center_code, center_name, occ_code, occ_name, level), data in sorted(formal_data.items()):
            level_prefix = get_level_prefix(level)
            formal_by_level[level_prefix].append({
                'center_code': center_code,
                'center_name': center_name,
                'occ_code': occ_code,
                'occ_name': occ_name,
                'level': level,  # Keep original level name for display
                'count': data['count'],
                'amount': data['amount']
            })
        
        # Create formal level sheets
        for level_prefix in formal_level_prefixes:
            if level_prefix in formal_by_level:
                ws = create_sheet_with_headers(wb, level_prefix, formal_headers)
                row = 2
                for item in formal_by_level[level_prefix]:
                    ws.cell(row=row, column=1, value=item['center_code'])
                    ws.cell(row=row, column=2, value=item['center_name'])
                    ws.cell(row=row, column=3, value=item['occ_code'])
                    ws.cell(row=row, column=4, value=item['occ_name'])
                    ws.cell(row=row, column=5, value='Formal')
                    ws.cell(row=row, column=6, value=item['level'])
                    ws.cell(row=row, column=7, value=item['count'])
                    ws.cell(row=row, column=8, value=float(item['amount']))
                    row += 1
        
        # Create Modular sheet (aggregated by center/occ to avoid duplication)
        if modular_data:
            ws = create_sheet_with_headers(wb, 'Modular', modular_headers)
            row = 2
            for (center_code, center_name, occ_code, occ_name), data in sorted(modular_data.items()):
                ws.cell(row=row, column=1, value=center_code)
                ws.cell(row=row, column=2, value=center_name)
                ws.cell(row=row, column=3, value=occ_code)
                ws.cell(row=row, column=4, value=occ_name)
                ws.cell(row=row, column=5, value='Modular')
                ws.cell(row=row, column=6, value=data['count'])
                ws.cell(row=row, column=7, value=float(data['amount']))
                row += 1
        
        # Create Workers PAS sheet
        if workers_pas_data:
            ws = create_sheet_with_headers(wb, 'Workers PAS', workers_pas_headers)
            row = 2
            for (center_code, center_name, occ_code, occ_name), data in sorted(workers_pas_data.items()):
                ws.cell(row=row, column=1, value=center_code)
                ws.cell(row=row, column=2, value=center_name)
                ws.cell(row=row, column=3, value=occ_code)
                ws.cell(row=row, column=4, value=occ_name)
                ws.cell(row=row, column=5, value="Worker's PAS")
                ws.cell(row=row, column=6, value=data['count'])
                ws.cell(row=row, column=7, value=float(data['amount']))
                row += 1
        
        # If no data at all, create an empty summary sheet
        if not wb.sheetnames:
            ws = wb.create_sheet(title="No Data")
            ws.cell(row=1, column=1, value="No enrollment data found for this series")

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_series_name = str(series.name).replace(" ", "_").lower()
        response['Content-Disposition'] = f'attachment; filename=registration_summary_{safe_series_name}_{date.today().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        
        return response
