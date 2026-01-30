from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from io import BytesIO

from candidates.models import EnrollmentModule, Candidate, CandidateEnrollment, EnrollmentPaper
from occupations.models import OccupationModule, OccupationLevel, OccupationPaper
from assessment_series.models import AssessmentSeries
from results.models import ModularResult


class MarksheetViewSet(viewsets.ViewSet):
    """
    ViewSet for generating marksheets
    """
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=['post'], url_path='generate-modular')
    def generate_modular_marksheet(self, request):
        """Generate Excel marksheet for modular candidates"""
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        module_id = request.data.get('module')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        
        if not all([assessment_series_id, occupation_id, module_id]):
            return Response(
                {'error': 'Assessment series, occupation, and module are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            module = OccupationModule.objects.get(id=module_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationModule.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or module'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get enrolled candidates for this module
        enrollments = EnrollmentModule.objects.filter(
            module=module,
            enrollment__assessment_series=assessment_series,
            enrollment__candidate__registration_category='modular'
        ).select_related(
            'enrollment__candidate',
            'enrollment__candidate__occupation'
        ).order_by('enrollment__candidate__registration_number')
        
        # Filter by assessment center if provided
        if assessment_center_id:
            enrollments = enrollments.filter(
                enrollment__candidate__assessment_center_id=assessment_center_id
            )
        
        if not enrollments.exists():
            return Response(
                {'error': 'No enrolled candidates found for the selected parameters'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Marksheet"
        
        # Define headers
        headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                  'CATEGORY', 'MODULE CODE', 'PRACTICAL']
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for idx, enrollment_module in enumerate(enrollments, 1):
            candidate = enrollment_module.enrollment.candidate
            
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=candidate.registration_number)
            ws.cell(row=idx+1, column=3, value=candidate.full_name)
            ws.cell(row=idx+1, column=4, value=candidate.occupation.occ_code if candidate.occupation else '')
            ws.cell(row=idx+1, column=5, value='Modular')
            ws.cell(row=idx+1, column=6, value=module.module_code)
            ws.cell(row=idx+1, column=7, value='')  # Empty for practical marks to be filled
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Marksheet_{module.module_code}_{assessment_series.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'], url_path='generate-formal')
    def generate_formal_marksheet(self, request):
        """Generate Excel marksheet for formal candidates"""
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        structure_type = request.data.get('structure_type')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        
        if not all([assessment_series_id, occupation_id, level_id, structure_type]):
            return Response(
                {'error': 'Assessment series, occupation, level, and structure type are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get formal candidates for this occupation and level through enrollments
        candidates = Candidate.objects.filter(
            occupation_id=occupation_id,
            registration_category='formal',
            enrollments__assessment_series=assessment_series,
            enrollments__occupation_level=level
        ).select_related('occupation', 'assessment_center').distinct().order_by('registration_number')
        
        # Filter by assessment center if provided
        if assessment_center_id:
            candidates = candidates.filter(assessment_center_id=assessment_center_id)
        
        if not candidates.exists():
            return Response(
                {'error': 'No candidates found for the selected parameters'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Marksheet"
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if structure_type == 'papers':
            # Paper-based marksheet
            # Get all papers for this level
            papers = OccupationPaper.objects.filter(level=level).order_by('paper_code')
            
            if not papers.exists():
                return Response(
                    {'error': 'No papers found for this level'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Define headers: Basic info + paper codes
            headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                      'CATEGORY', 'LEVEL']
            
            # Add paper codes as columns
            for paper in papers:
                headers.append(paper.paper_code)
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Write data rows
            for idx, candidate in enumerate(candidates, 1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=candidate.registration_number)
                ws.cell(row=idx+1, column=3, value=candidate.full_name)
                ws.cell(row=idx+1, column=4, value=candidate.occupation.occ_code if candidate.occupation else '')
                ws.cell(row=idx+1, column=5, value='Formal')
                ws.cell(row=idx+1, column=6, value=level.level_name)
                
                # Empty cells for paper marks
                for paper_idx in range(len(papers)):
                    ws.cell(row=idx+1, column=7+paper_idx, value='')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            
            # Set width for paper columns
            for i in range(len(papers)):
                col_letter = get_column_letter(7 + i)
                ws.column_dimensions[col_letter].width = 12
        
        else:  # structure_type == 'modules'
            # Module-based marksheet with Theory and Practical columns
            # Note: "Module-based" here means the level is assessed with Theory + Practical
            # This is NOT related to OccupationModule (which is for modular candidates only)
            
            # Define headers: Basic info + Theory/Practical columns
            headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                      'CATEGORY', 'LEVEL', 'THEORY', 'PRACTICAL']
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Write data rows
            for idx, candidate in enumerate(candidates, 1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=candidate.registration_number)
                ws.cell(row=idx+1, column=3, value=candidate.full_name)
                ws.cell(row=idx+1, column=4, value=candidate.occupation.occ_code if candidate.occupation else '')
                ws.cell(row=idx+1, column=5, value='Formal')
                ws.cell(row=idx+1, column=6, value=level.level_name)
                ws.cell(row=idx+1, column=7, value='')  # Theory
                ws.cell(row=idx+1, column=8, value='')  # Practical
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Marksheet_Formal_{level.level_name.replace(' ', '_')}_{assessment_series.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'], url_path='generate-workers-pas')
    def generate_workers_pas_marksheet(self, request):
        """Generate Excel marksheet for Worker's PAS candidates with highlighted enrolled papers"""
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        
        if not all([assessment_series_id, occupation_id, level_id]):
            return Response(
                {'error': 'Assessment series, occupation, and level are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all papers for this level
        papers = OccupationPaper.objects.filter(level=level).order_by('paper_code')
        
        if not papers.exists():
            return Response(
                {'error': 'No papers found for this level'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get Worker's PAS candidates with enrollments
        # Note: Worker's PAS can select papers from any level, but we filter by papers in the selected level
        enrollments = CandidateEnrollment.objects.filter(
            assessment_series=assessment_series,
            candidate__occupation_id=occupation_id,
            candidate__registration_category='workers_pas',
            papers__paper__level=level
        ).select_related('candidate', 'candidate__occupation', 'candidate__assessment_center').prefetch_related(
            'papers__paper'
        ).distinct().order_by('candidate__registration_number')
        
        # Filter by assessment center if provided
        if assessment_center_id:
            enrollments = enrollments.filter(candidate__assessment_center_id=assessment_center_id)
        
        if not enrollments.exists():
            return Response(
                {'error': 'No candidates found for the selected parameters'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Marksheet"
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Style for highlighted cells (yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Define headers: Basic info + paper codes
        headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                  'CATEGORY', 'LEVEL']
        
        # Add paper codes as columns
        paper_list = list(papers)
        for paper in paper_list:
            headers.append(paper.paper_code)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for idx, enrollment in enumerate(enrollments, 1):
            candidate = enrollment.candidate
            
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=candidate.registration_number)
            ws.cell(row=idx+1, column=3, value=candidate.full_name)
            ws.cell(row=idx+1, column=4, value=candidate.occupation.occ_code if candidate.occupation else '')
            ws.cell(row=idx+1, column=5, value="Worker's PAS")
            ws.cell(row=idx+1, column=6, value=level.level_name)
            
            # Get enrolled papers for this candidate
            enrolled_paper_ids = set(enrollment.papers.values_list('paper_id', flat=True))
            
            # Fill paper columns and highlight enrolled ones
            for paper_idx, paper in enumerate(paper_list):
                cell = ws.cell(row=idx+1, column=7+paper_idx, value='')
                
                # Highlight if candidate is enrolled in this paper
                if paper.id in enrolled_paper_ids:
                    cell.fill = highlight_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Set width for paper columns
        for i in range(len(paper_list)):
            col_letter = get_column_letter(7 + i)
            ws.column_dimensions[col_letter].width = 12
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Marksheet_WorkersPAS_{level.level_name.replace(' ', '_')}_{assessment_series.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'], url_path='upload-modular', parser_classes=[MultiPartParser, FormParser])
    def upload_modular_marks(self, request):
        """Upload marks from Excel file for modular candidates"""
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        module_id = request.data.get('module')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        excel_file = request.FILES.get('file')
        
        if not all([assessment_series_id, occupation_id, module_id, excel_file]):
            return Response(
                {'error': 'Assessment series, occupation, module, and Excel file are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            module = OccupationModule.objects.get(id=module_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationModule.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or module'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Parse Excel file
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate headers
        expected_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                          'CATEGORY', 'MODULE CODE', 'PRACTICAL']
        actual_headers = [cell.value for cell in ws[1]]
        
        if actual_headers[:7] != expected_headers:
            return Response(
                {'error': 'Invalid Excel format. Please use the generated marksheet template.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Process rows and collect results
        errors = []
        updated_count = 0
        skipped_count = 0
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[1]:  # Skip empty rows
                    continue
                
                reg_number = str(row[1]).strip()
                module_code = str(row[5]).strip() if row[5] else ''
                practical_mark = row[6]
                
                # Validate module code matches
                if module_code != module.module_code:
                    errors.append(f'Row {row_num}: Module code mismatch. Expected {module.module_code}, got {module_code}')
                    continue
                
                # Validate practical mark
                if practical_mark is None or practical_mark == '':
                    skipped_count += 1
                    continue
                
                try:
                    practical_mark = float(practical_mark)
                    if practical_mark < 0 or practical_mark > 100:
                        errors.append(f'Row {row_num}: Invalid mark {practical_mark}. Must be between 0 and 100')
                        continue
                except (ValueError, TypeError):
                    errors.append(f'Row {row_num}: Invalid mark value "{practical_mark}"')
                    continue
                
                # Find candidate
                try:
                    candidate = Candidate.objects.get(
                        registration_number=reg_number,
                        registration_category='modular'
                    )
                except Candidate.DoesNotExist:
                    errors.append(f'Row {row_num}: Candidate {reg_number} not found')
                    continue
                
                # Find enrollment
                try:
                    enrollment = CandidateEnrollment.objects.get(
                        candidate=candidate,
                        assessment_series=assessment_series
                    )
                except CandidateEnrollment.DoesNotExist:
                    errors.append(f'Row {row_num}: No enrollment found for {reg_number} in this series')
                    continue
                
                # Find enrollment module
                try:
                    enrollment_module = EnrollmentModule.objects.get(
                        enrollment=enrollment,
                        module=module
                    )
                except EnrollmentModule.DoesNotExist:
                    errors.append(f'Row {row_num}: Candidate {reg_number} not enrolled in module {module_code}')
                    continue
                
                # Filter by assessment center if provided
                if assessment_center_id and str(candidate.assessment_center_id) != str(assessment_center_id):
                    skipped_count += 1
                    continue
                
                # Create or update result
                result, created = ModularResult.objects.update_or_create(
                    candidate=candidate,
                    assessment_series=assessment_series,
                    module=module,
                    type='practical',  # Modular is practical
                    defaults={
                        'mark': practical_mark,
                        'status': 'normal'
                    }
                )
                updated_count += 1
        
        # Prepare response
        response_data = {
            'message': f'Marks successfully uploaded. {updated_count} candidate(s) updated.',
            'updated_count': updated_count,
            'skipped_count': skipped_count,
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' {len(errors)} error(s) encountered.'
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='upload-formal', parser_classes=[MultiPartParser, FormParser])
    def upload_formal_marks(self, request):
        """Upload marks from Excel file for formal candidates (module-based or paper-based)"""
        from results.models import FormalResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        excel_file = request.FILES.get('file')
        
        if not all([assessment_series_id, occupation_id, level_id, excel_file]):
            return Response(
                {'error': 'Assessment series, occupation, level, and Excel file are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Parse Excel file
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get headers to determine structure type
        headers = [cell.value for cell in ws[1]]
        
        # Determine if module-based or paper-based
        is_module_based = 'THEORY' in headers and 'PRACTICAL' in headers
        is_paper_based = not is_module_based
        
        errors = []
        updated_count = 0
        skipped_count = 0
        
        if is_module_based:
            # Module-based: THEORY and PRACTICAL columns
            expected_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                              'CATEGORY', 'LEVEL', 'THEORY', 'PRACTICAL']
            
            if headers[:8] != expected_headers:
                return Response(
                    {'error': 'Invalid Excel format for module-based. Please use the generated marksheet template.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[1]:  # Skip empty rows
                        continue
                    
                    reg_number = str(row[1]).strip()
                    theory_mark = row[6]
                    practical_mark = row[7]
                    
                    # Skip if both marks are empty
                    if (theory_mark is None or theory_mark == '') and (practical_mark is None or practical_mark == ''):
                        skipped_count += 1
                        continue
                    
                    # Find candidate
                    try:
                        candidate = Candidate.objects.get(
                            registration_number=reg_number,
                            registration_category='formal'
                        )
                    except Candidate.DoesNotExist:
                        errors.append(f'Row {row_num}: Candidate {reg_number} not found')
                        continue
                    
                    # Verify enrollment
                    enrollment_exists = CandidateEnrollment.objects.filter(
                        candidate=candidate,
                        assessment_series=assessment_series,
                        occupation_level=level
                    ).exists()
                    
                    if not enrollment_exists:
                        errors.append(f'Row {row_num}: Candidate {reg_number} not enrolled in this level')
                        continue
                    
                    # Filter by assessment center if provided
                    if assessment_center_id and str(candidate.assessment_center_id) != str(assessment_center_id):
                        skipped_count += 1
                        continue
                    
                    # Update theory result if provided
                    if theory_mark is not None and theory_mark != '':
                        try:
                            theory_mark = float(theory_mark)
                            if theory_mark < 0 or theory_mark > 100:
                                errors.append(f'Row {row_num}: Invalid theory mark {theory_mark}. Must be between 0 and 100')
                            else:
                                FormalResult.objects.update_or_create(
                                    candidate=candidate,
                                    assessment_series=assessment_series,
                                    level=level,
                                    type='theory',
                                    defaults={
                                        'mark': theory_mark,
                                        'status': 'normal'
                                    }
                                )
                                updated_count += 1
                        except (ValueError, TypeError):
                            errors.append(f'Row {row_num}: Invalid theory mark value "{theory_mark}"')
                    
                    # Update practical result if provided
                    if practical_mark is not None and practical_mark != '':
                        try:
                            practical_mark = float(practical_mark)
                            if practical_mark < 0 or practical_mark > 100:
                                errors.append(f'Row {row_num}: Invalid practical mark {practical_mark}. Must be between 0 and 100')
                            else:
                                FormalResult.objects.update_or_create(
                                    candidate=candidate,
                                    assessment_series=assessment_series,
                                    level=level,
                                    type='practical',
                                    defaults={
                                        'mark': practical_mark,
                                        'status': 'normal'
                                    }
                                )
                                updated_count += 1
                        except (ValueError, TypeError):
                            errors.append(f'Row {row_num}: Invalid practical mark value "{practical_mark}"')
        
        else:
            # Paper-based: Individual paper columns
            # Get papers for this level
            papers = OccupationPaper.objects.filter(level=level).order_by('paper_code')
            
            if not papers.exists():
                return Response(
                    {'error': 'No papers found for this level'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Validate headers
            expected_base_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                                    'CATEGORY', 'LEVEL']
            
            if headers[:6] != expected_base_headers:
                return Response(
                    {'error': 'Invalid Excel format for paper-based. Please use the generated marksheet template.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Map paper codes to column indices
            paper_columns = {}
            for idx, header in enumerate(headers[6:], start=6):
                if header:
                    paper_columns[header] = idx
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[1]:  # Skip empty rows
                        continue
                    
                    reg_number = str(row[1]).strip()
                    
                    # Find candidate
                    try:
                        candidate = Candidate.objects.get(
                            registration_number=reg_number,
                            registration_category='formal'
                        )
                    except Candidate.DoesNotExist:
                        errors.append(f'Row {row_num}: Candidate {reg_number} not found')
                        continue
                    
                    # Verify enrollment
                    enrollment_exists = CandidateEnrollment.objects.filter(
                        candidate=candidate,
                        assessment_series=assessment_series,
                        occupation_level=level
                    ).exists()
                    
                    if not enrollment_exists:
                        errors.append(f'Row {row_num}: Candidate {reg_number} not enrolled in this level')
                        continue
                    
                    # Filter by assessment center if provided
                    if assessment_center_id and str(candidate.assessment_center_id) != str(assessment_center_id):
                        skipped_count += 1
                        continue
                    
                    # Process each paper
                    row_updated = False
                    for paper in papers:
                        col_idx = paper_columns.get(paper.paper_code)
                        if col_idx is None:
                            continue
                        
                        mark = row[col_idx] if col_idx < len(row) else None
                        
                        if mark is None or mark == '':
                            continue
                        
                        try:
                            mark = float(mark)
                            if mark < 0 or mark > 100:
                                errors.append(f'Row {row_num}: Invalid mark {mark} for paper {paper.paper_code}. Must be between 0 and 100')
                                continue
                            
                            FormalResult.objects.update_or_create(
                                candidate=candidate,
                                assessment_series=assessment_series,
                                level=level,
                                paper=paper,
                                type=paper.paper_type,  # Use the paper's actual type (theory/practical)
                                defaults={
                                    'mark': mark,
                                    'status': 'normal'
                                }
                            )
                            row_updated = True
                        except (ValueError, TypeError):
                            errors.append(f'Row {row_num}: Invalid mark value "{mark}" for paper {paper.paper_code}')
                    
                    if row_updated:
                        updated_count += 1
        
        # Prepare response
        response_data = {
            'message': f'Marks successfully uploaded. {updated_count} result(s) updated.',
            'updated_count': updated_count,
            'skipped_count': skipped_count,
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' {len(errors)} error(s) encountered.'
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='upload-workers-pas', parser_classes=[MultiPartParser, FormParser])
    def upload_workers_pas_marks(self, request):
        """Upload marks from Excel file for Worker's PAS candidates"""
        from results.models import WorkersPasResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        assessment_center_id = request.data.get('assessment_center')  # Optional
        excel_file = request.FILES.get('file')
        
        if not all([assessment_series_id, occupation_id, level_id, excel_file]):
            return Response(
                {'error': 'Assessment series, occupation, level, and Excel file are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Parse Excel file
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        # Validate base headers
        expected_base_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 
                                'CATEGORY', 'LEVEL']
        
        if headers[:6] != expected_base_headers:
            return Response(
                {'error': 'Invalid Excel format. Please use the generated Worker\'s PAS marksheet template.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get papers for this level
        papers = OccupationPaper.objects.filter(level=level).order_by('paper_code')
        
        if not papers.exists():
            return Response(
                {'error': 'No papers found for this level'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Map paper codes to column indices
        paper_columns = {}
        for idx, header in enumerate(headers[6:], start=6):
            if header:
                paper_columns[header] = idx
        
        errors = []
        updated_count = 0
        skipped_count = 0
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[1]:  # Skip empty rows
                    continue
                
                reg_number = str(row[1]).strip()
                
                # Find candidate
                try:
                    candidate = Candidate.objects.get(
                        registration_number=reg_number,
                        registration_category='workers_pas'
                    )
                except Candidate.DoesNotExist:
                    errors.append(f'Row {row_num}: Candidate {reg_number} not found')
                    continue
                
                # Verify enrollment - Worker's PAS enrollments don't have occupation_level set
                # Instead, check if they have papers from this level
                enrollment = CandidateEnrollment.objects.filter(
                    candidate=candidate,
                    assessment_series=assessment_series
                ).first()
                
                if not enrollment:
                    errors.append(f'Row {row_num}: Candidate {reg_number} not enrolled in {assessment_series.name}')
                    continue
                
                # Get enrolled papers for this candidate from the specified level
                enrolled_papers = EnrollmentPaper.objects.filter(
                    enrollment=enrollment,
                    paper__level=level
                ).values_list('paper_id', flat=True)
                
                if not enrolled_papers:
                    errors.append(f'Row {row_num}: Candidate {reg_number} has no papers enrolled in {level.level_name}')
                    continue
                
                # Filter by assessment center if provided
                if assessment_center_id and str(candidate.assessment_center_id) != str(assessment_center_id):
                    skipped_count += 1
                    continue
                
                # Process each paper
                row_updated = False
                for paper in papers:
                    # Only process papers the candidate is enrolled in
                    if paper.id not in enrolled_papers:
                        continue
                    
                    col_idx = paper_columns.get(paper.paper_code)
                    if col_idx is None:
                        continue
                    
                    mark = row[col_idx] if col_idx < len(row) else None
                    
                    if mark is None or mark == '':
                        continue
                    
                    try:
                        mark = float(mark)
                        if mark < 0 or mark > 100:
                            errors.append(f'Row {row_num}: Invalid mark {mark} for paper {paper.paper_code}. Must be between 0 and 100')
                            continue
                        
                        # Get module for this paper
                        module = paper.module
                        if not module:
                            errors.append(f'Row {row_num}: Paper {paper.paper_code} has no associated module')
                            continue
                        
                        WorkersPasResult.objects.update_or_create(
                            candidate=candidate,
                            assessment_series=assessment_series,
                            level=level,
                            module=module,
                            paper=paper,
                            defaults={
                                'mark': mark,
                                'status': 'normal'
                            }
                        )
                        row_updated = True
                    except (ValueError, TypeError):
                        errors.append(f'Row {row_num}: Invalid mark value "{mark}" for paper {paper.paper_code}')
                
                if row_updated:
                    updated_count += 1
                elif not any(f'Row {row_num}:' in err for err in errors):
                    skipped_count += 1
        
        # Prepare response
        response_data = {
            'message': f'✓ Successfully uploaded marks for {updated_count} candidate(s).',
            'updated_count': updated_count,
            'skipped_count': skipped_count,
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] = f'⚠ Upload completed with warnings. {updated_count} candidate(s) updated, {len(errors)} error(s) encountered.'
        
        if updated_count == 0 and len(errors) > 0:
            response_data['message'] = f'✗ Upload failed. No candidates were updated due to {len(errors)} error(s).'
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(response_data, status=status.HTTP_200_OK)

    def _header_footer(self, canvas, doc, title_text, subtitle_text):
        """Draw header and footer on each page"""
        canvas.saveState()
        
        # Header
        # Logo
        try:
            # Assuming logo is in backend/static/images/uvtab-logo.png
            # Adjust path relative to where manage.py is run or use absolute path
            import os
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, 'backend/static/images/uvtab-logo.png')
            if not os.path.exists(logo_path):
                 # Try alternative path if main one fails (dev environment structure vary)
                 logo_path = os.path.join(settings.BASE_DIR, 'static/images/uvtab-logo.png')
            
            if os.path.exists(logo_path):
                canvas.drawImage(logo_path, 30, 750, width=60, height=60, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass # Skip logo if not found
            
        # Header Text (Centered)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(300, 790, "UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(300, 775, "P.O. Box 1499, Kampala. Plot 891, Kigobe Road, Kyambogo Hill, Kampala, Uganda")
        canvas.drawCentredString(300, 762, "Tel: +256 392-002468 | Email: info@uvtab.go.ug")
        
        # Title and Subtitle
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(300, 730, title_text)
        canvas.setFont("Helvetica-Oblique", 11)
        canvas.drawCentredString(300, 715, subtitle_text)
        
        # Line below header
        canvas.line(30, 705, 565, 705)
        
        # Footer
        canvas.line(30, 50, 565, 50)
        canvas.setFont("Helvetica", 9)
        canvas.drawString(30, 35, "Generated from EMIS System")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(565, 35, f"Page {page_num}")
        
        canvas.restoreState()

    @action(detail=False, methods=['post'], url_path='print-modular')
    def print_modular_marksheet(self, request):
        """Generate PDF marksheet for modular candidates"""
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        module_id = request.data.get('module')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, module_id]):
            return Response(
                {'error': 'Assessment series, occupation, and module are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            module = OccupationModule.objects.get(id=module_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationModule.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or module'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        enrollments = EnrollmentModule.objects.filter(
            module=module,
            enrollment__assessment_series=assessment_series,
            enrollment__candidate__registration_category='modular'
        ).select_related(
            'enrollment__candidate',
            'enrollment__candidate__occupation'
        ).order_by('enrollment__candidate__registration_number')
        
        if assessment_center_id:
            enrollments = enrollments.filter(
                enrollment__candidate__assessment_center_id=assessment_center_id
            )
            
        if not enrollments.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )
            
        response = HttpResponse(content_type='application/pdf')
        filename = f"Marksheet_{module.module_code}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        buffer = BytesIO()
        # Use Portrait A4
        from reportlab.lib.pagesizes import portrait
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), topMargin=150, bottomMargin=60, title=f"Marksheet - {module.module_code} {module.module_name}")
        elements = []
        styles = getSampleStyleSheet()
        normal_style = styles['Normal']
        
        # Table Data
        # Columns: SN, Reg No, Name, Mark
        # Reduced columns to fit portrait and ensure wrapping
        data = [[
            Paragraph('<b>SN</b>', normal_style), 
            Paragraph('<b>Reg No</b>', normal_style), 
            Paragraph('<b>Name</b>', normal_style), 
            Paragraph('<b>Mark</b>', normal_style)
        ]]
        
        for idx, em in enumerate(enrollments, 1):
            cand = em.enrollment.candidate
            try:
                result = ModularResult.objects.get(
                    candidate=cand,
                    assessment_series=assessment_series,
                    module=module
                )
                mark = str(result.mark if result.mark is not None else '')
            except ModularResult.DoesNotExist:
                mark = ''
                
            data.append([
                str(idx),
                Paragraph(cand.registration_number, normal_style),
                Paragraph(cand.full_name, normal_style),
                mark
            ])
            
        # Adjust column widths for Portrait (A4 width ~595pt, margins ~60pt/side -> ~475pt usable)
        # 30 + 130 + 220 + 80 = 460
        # Increased Reg No width to 130
        table = Table(data, colWidths=[30, 130, 220, 80], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'), # Center SN
            ('ALIGN', (-1, 0), (-1, -1), 'CENTER'), # Center Marks
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Title Texts for Header
        title_text = f"Marksheet - {module.module_code} {module.module_name}"
        subtitle_text = f"Series: {assessment_series.name} | Category: Modular"
        
        doc.build(elements, onFirstPage=lambda c, d: self._header_footer(c, d, title_text, subtitle_text),
                  onLaterPages=lambda c, d: self._header_footer(c, d, title_text, subtitle_text))
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    @action(detail=False, methods=['post'], url_path='print-formal')
    def print_formal_marksheet(self, request):
        """Generate PDF marksheet for formal candidates"""
        from results.models import FormalResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        structure_type = request.data.get('structure_type')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, level_id]):
            return Response(
                {'error': 'Assessment series, occupation, and level are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
            occupation = level.occupation
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        candidates = Candidate.objects.filter(
            occupation_id=occupation_id,
            registration_category='formal',
            enrollments__assessment_series=assessment_series,
            enrollments__occupation_level=level
        ).select_related('occupation', 'assessment_center').distinct().order_by('registration_number')
        
        if assessment_center_id:
            candidates = candidates.filter(assessment_center_id=assessment_center_id)
            
        if not candidates.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )
            
        response = HttpResponse(content_type='application/pdf')
        filename = f"Marksheet_Formal_{level.level_name}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        buffer = BytesIO()
        from reportlab.lib.pagesizes import portrait
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), topMargin=150, bottomMargin=60, title=f"Marksheet - {level.level_name}")
        elements = []
        styles = getSampleStyleSheet()
        normal_style = styles['Normal']
        
        title_text = f"Marksheet - {level.level_name}"
        subtitle_text = f"Series: {assessment_series.name} | Category: Formal | Occupation: {occupation.occ_name} ({occupation.occ_code})"
        
        if structure_type == 'papers':
            papers = list(OccupationPaper.objects.filter(level=level).order_by('paper_code'))
            # Calculate dynamic column width
            # Usable width ~480. 
            # Fixed sizes: SN(30), Reg(130), Name(150) = 310. Remaining 170.
            # Increased Reg No from 90 to 130
            num_papers = len(papers)
            paper_col_width = max(40, 170 // max(1, num_papers))
            
            headers = [
                Paragraph('<b>SN</b>', normal_style),
                Paragraph('<b>Reg No</b>', normal_style),
                Paragraph('<b>Name</b>', normal_style)
            ] + [Paragraph(f'<b>{p.paper_code}</b>', normal_style) for p in papers]
            
            data = [headers]
            col_widths = [30, 130, 150] + [paper_col_width] * num_papers
            
            for idx, cand in enumerate(candidates, 1):
                row = [
                    str(idx),
                    Paragraph(cand.registration_number, normal_style),
                    Paragraph(cand.full_name, normal_style)
                ]
                for paper in papers:
                    try:
                        result = FormalResult.objects.get(
                            candidate=cand,
                            assessment_series=assessment_series,
                            level=level,
                            paper=paper
                        )
                        row.append(str(result.mark) if result.mark is not None else '')
                    except FormalResult.DoesNotExist:
                        row.append('')
                data.append(row)
            
            # Add table
            # Table generation moved to common block to avoid duplication
            
            # Prepare Key Table to be appended after the main table
            extra_elements = [Spacer(1, 20)]

            # Paper Key Table
            key_data = [[Paragraph('<b>Code</b>', normal_style), Paragraph('<b>Paper Name</b>', normal_style)]]
            for p in papers:
                 key_data.append([p.paper_code, Paragraph(p.paper_name, normal_style)])
            
            key_table = Table(key_data, colWidths=[80, 400])
            key_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            extra_elements.append(Paragraph("<b>Paper Codes Description:</b>", normal_style))
            extra_elements.append(Spacer(1, 5))
            extra_elements.append(key_table)
        else:
            extra_elements = []
            # Theory + Practical
            headers = [
                Paragraph('<b>SN</b>', normal_style),
                Paragraph('<b>Reg No</b>', normal_style),
                Paragraph('<b>Name</b>', normal_style),
                Paragraph('<b>Theory</b>', normal_style),
                Paragraph('<b>Practical</b>', normal_style)
            ]
            data = [headers]
            col_widths = [30, 100, 200, 75, 75]
            
            for idx, cand in enumerate(candidates, 1):
                row = [
                    str(idx),
                    Paragraph(cand.registration_number, normal_style),
                    Paragraph(cand.full_name, normal_style)
                ]
                
                # Theory
                try:
                    res_t = FormalResult.objects.get(candidate=cand, assessment_series=assessment_series, level=level, type='theory')
                    row.append(str(res_t.mark) if res_t.mark is not None else '')
                except FormalResult.DoesNotExist:
                    row.append('')
                    
                # Practical
                try:
                    res_p = FormalResult.objects.get(candidate=cand, assessment_series=assessment_series, level=level, type='practical')
                    row.append(str(res_p.mark) if res_p.mark is not None else '')
                except FormalResult.DoesNotExist:
                    row.append('')
                    
                data.append(row)
                
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'), # Center marks columns
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        doc.build(elements, onFirstPage=lambda c, d: self._header_footer(c, d, title_text, subtitle_text),
                  onLaterPages=lambda c, d: self._header_footer(c, d, title_text, subtitle_text))
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    @action(detail=False, methods=['post'], url_path='export-modular')
    def export_modular_results(self, request):
        """Export modular results to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        module_id = request.data.get('module')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, module_id]):
            return Response(
                {'error': 'Assessment series, occupation, and module are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            module = OccupationModule.objects.get(id=module_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationModule.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or module'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        enrollments = EnrollmentModule.objects.filter(
            module=module,
            enrollment__assessment_series=assessment_series,
            enrollment__candidate__registration_category='modular'
        ).select_related(
            'enrollment__candidate',
            'enrollment__candidate__occupation'
        ).order_by('enrollment__candidate__registration_number')
        
        if assessment_center_id:
            enrollments = enrollments.filter(
                enrollment__candidate__assessment_center_id=assessment_center_id
            )
            
        if not enrollments.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Headers
        headers = ['SN', 'Reg No', 'Name', 'Mark']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
        for idx, em in enumerate(enrollments, 1):
            cand = em.enrollment.candidate
            try:
                result = ModularResult.objects.get(
                    candidate=cand,
                    assessment_series=assessment_series,
                    module=module
                )
                mark = result.mark if result.mark is not None else ''
            except ModularResult.DoesNotExist:
                mark = ''
                
            ws.append([idx, cand.registration_number, cand.full_name, mark])
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Results_{module.module_code}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @action(detail=False, methods=['post'], url_path='export-formal')
    def export_formal_results(self, request):
        """Export formal results to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from results.models import FormalResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        structure_type = request.data.get('structure_type')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, level_id]):
            return Response(
                {'error': 'Assessment series, occupation, and level are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        candidates = Candidate.objects.filter(
            occupation_id=occupation_id,
            registration_category='formal',
            enrollments__assessment_series=assessment_series,
            enrollments__occupation_level=level
        ).select_related('occupation', 'assessment_center').distinct().order_by('registration_number')
        
        if assessment_center_id:
            candidates = candidates.filter(assessment_center_id=assessment_center_id)
            
        if not candidates.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        if structure_type == 'papers':
            papers = list(OccupationPaper.objects.filter(level=level).order_by('paper_code'))
            headers = ['SN', 'Reg No', 'Name'] + [p.paper_code for p in papers]
            ws.append(headers)
            
            for idx, cand in enumerate(candidates, 1):
                row = [idx, cand.registration_number, cand.full_name]
                for paper in papers:
                    try:
                        result = FormalResult.objects.get(
                            candidate=cand,
                            assessment_series=assessment_series,
                            level=level,
                            paper=paper
                        )
                        row.append(result.mark if result.mark is not None else '')
                    except FormalResult.DoesNotExist:
                        row.append('')
                ws.append(row)
        else:
            headers = ['SN', 'Reg No', 'Name', 'Theory', 'Practical']
            ws.append(headers)
            
            for idx, cand in enumerate(candidates, 1):
                row = [idx, cand.registration_number, cand.full_name]
                # Theory
                try:
                    res_t = FormalResult.objects.get(candidate=cand, assessment_series=assessment_series, level=level, type='theory')
                    row.append(res_t.mark if res_t.mark is not None else '')
                except FormalResult.DoesNotExist:
                    row.append('')
                # Practical
                try:
                    res_p = FormalResult.objects.get(candidate=cand, assessment_series=assessment_series, level=level, type='practical')
                    row.append(res_p.mark if res_p.mark is not None else '')
                except FormalResult.DoesNotExist:
                    row.append('')
                ws.append(row)

        # Append Paper Key for Paper-based structure
        if structure_type == 'papers':
            ws.append([])  # Empty row
            ws.append([])  # Empty row
            ws.append(['Paper Codes Description:'])
            ws['A' + str(ws.max_row)].font = Font(bold=True)
            
            ws.append(['Code', 'Paper Name'])
            # Style the key header
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style='thin'))
                
            for p in papers:
                ws.append([p.paper_code, p.paper_name])

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Results_Formal_{level.level_name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @action(detail=False, methods=['post'], url_path='export-workers-pas')
    def export_workers_pas_results(self, request):
        """Export Workers PAS results to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from results.models import WorkersPasResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, level_id]):
            return Response(
                {'error': 'Assessment series, occupation, and level are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        papers = list(OccupationPaper.objects.filter(level=level).order_by('paper_code'))
            
        enrollments = CandidateEnrollment.objects.filter(
            assessment_series=assessment_series,
            candidate__occupation_id=occupation_id,
            candidate__registration_category='workers_pas',
            papers__paper__level=level
        ).select_related('candidate', 'candidate__occupation').prefetch_related(
            'papers__paper'
        ).distinct().order_by('candidate__registration_number')
        
        if assessment_center_id:
            enrollments = enrollments.filter(candidate__assessment_center_id=assessment_center_id)
            
        if not enrollments.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        headers = ['SN', 'Reg No', 'Name'] + [p.paper_code for p in papers]
        ws.append(headers)
        
        for idx, enrollment in enumerate(enrollments, 1):
            cand = enrollment.candidate
            row = [idx, cand.registration_number, cand.full_name]
            enrolled_paper_ids = set(enrollment.papers.values_list('paper_id', flat=True))
            
            for paper in papers:
                if paper.id not in enrolled_paper_ids:
                    row.append('N/A')
                else:
                    try:
                        result = WorkersPasResult.objects.get(
                            candidate=cand,
                            assessment_series=assessment_series,
                            paper=paper
                        )
                        row.append(result.mark if result.mark is not None else '')
                    except WorkersPasResult.DoesNotExist:
                        row.append('')
            ws.append(row)

        # Append Paper Key for Workers PAS
        ws.append([])  # Empty row
        ws.append([])  # Empty row
        ws.append(['Paper Codes Description:'])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        
        ws.append(['Code', 'Paper Name'])
        # Style the key header
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            
        for p in papers:
            ws.append([p.paper_code, p.paper_name])

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Results_WorkersPAS_{level.level_name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @action(detail=False, methods=['post'], url_path='print-workers-pas')
    def print_workers_pas_marksheet(self, request):
        """Generate PDF marksheet for Workers PAS candidates"""
        from results.models import WorkersPasResult
        
        assessment_series_id = request.data.get('assessment_series')
        occupation_id = request.data.get('occupation')
        level_id = request.data.get('level')
        assessment_center_id = request.data.get('assessment_center')
        
        if not all([assessment_series_id, occupation_id, level_id]):
            return Response(
                {'error': 'Assessment series, occupation, and level are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
            level = OccupationLevel.objects.get(id=level_id, occupation_id=occupation_id)
            occupation = level.occupation
        except (AssessmentSeries.DoesNotExist, OccupationLevel.DoesNotExist):
            return Response(
                {'error': 'Invalid assessment series or level'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        papers = list(OccupationPaper.objects.filter(level=level).order_by('paper_code'))
            
        enrollments = CandidateEnrollment.objects.filter(
            assessment_series=assessment_series,
            candidate__occupation_id=occupation_id,
            candidate__registration_category='workers_pas',
            papers__paper__level=level
        ).select_related('candidate', 'candidate__occupation').prefetch_related(
            'papers__paper'
        ).distinct().order_by('candidate__registration_number')
        
        if assessment_center_id:
            enrollments = enrollments.filter(candidate__assessment_center_id=assessment_center_id)
            
        if not enrollments.exists():
            return Response(
                {'error': 'No candidates found'},
                status=status.HTTP_404_NOT_FOUND
            )
            
        response = HttpResponse(content_type='application/pdf')
        filename = f"Marksheet_WorkersPAS_{level.level_name}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        buffer = BytesIO()
        from reportlab.lib.pagesizes import portrait
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), topMargin=150, bottomMargin=60, title=f"Marksheet - Worker's PAS - {level.level_name}")
        elements = []
        styles = getSampleStyleSheet()
        normal_style = styles['Normal']
        
        title_text = f"Marksheet - Worker's PAS - {level.level_name}"
        subtitle_text = f"Series: {assessment_series.name} | Occupation: {occupation.occ_name}"
        
        # Calculate dynamic column width
        # Usable width ~480. 
        # Fixed: SN(30), Reg(130), Name(150) = 310. Remaining 170.
        num_papers = len(papers)
        paper_col_width = max(40, 170 // max(1, num_papers))
        
        headers = [
            Paragraph('<b>SN</b>', normal_style),
            Paragraph('<b>Reg No</b>', normal_style),
            Paragraph('<b>Name</b>', normal_style)
        ] + [Paragraph(f'<b>{p.paper_code}</b>', normal_style) for p in papers]
        
        data = [headers]
        col_widths = [30, 130, 150] + [paper_col_width] * num_papers
        
        for idx, enrollment in enumerate(enrollments, 1):
            cand = enrollment.candidate
            row = [
                str(idx),
                Paragraph(cand.registration_number, normal_style),
                Paragraph(cand.full_name, normal_style)
            ]
            
            enrolled_paper_ids = set(enrollment.papers.values_list('paper_id', flat=True))
            
            for paper in papers:
                if paper.id not in enrolled_paper_ids:
                    row.append('N/A')
                else:
                    try:
                        result = WorkersPasResult.objects.get(
                            candidate=cand,
                            assessment_series=assessment_series,
                            paper=paper
                        )
                        row.append(str(result.mark) if result.mark is not None else '')
                    except WorkersPasResult.DoesNotExist:
                        row.append('')
            data.append(row)
            
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Paper Key Table
        key_data = [[Paragraph('<b>Code</b>', normal_style), Paragraph('<b>Paper Name</b>', normal_style)]]
        for p in papers:
             key_data.append([p.paper_code, Paragraph(p.paper_name, normal_style)])
        
        key_table = Table(key_data, colWidths=[80, 400])
        key_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(Paragraph("<b>Paper Codes Description:</b>", normal_style))
        elements.append(Spacer(1, 5))
        elements.append(key_table)
        
        doc.build(elements, onFirstPage=lambda c, d: self._header_footer(c, d, title_text, subtitle_text),
                  onLaterPages=lambda c, d: self._header_footer(c, d, title_text, subtitle_text))
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response


def calculate_grade(mark, grade_type='practical'):
    """
    Calculate grade based on mark using UVTAB grading system
    
    Practical grading (pass mark: 65%):
    A+: 90-100, A: 85-89, B+: 75-84, B: 65-74, B-: 60-64, 
    C: 55-59, C-: 50-54, D: 40-49, D-: 30-39, E: 0-29
    
    Theory grading (pass mark: 50%):
    A+: 85-100, A: 80-84, B: 70-79, B-: 60-69, C: 50-59,
    C-: 40-49, D: 30-39, E: 0-29
    """
    if mark is None or mark == -1:
        return None
    
    if grade_type == 'practical':
        # Practical grading
        if mark >= 90:
            return 'A+'
        elif mark >= 85:
            return 'A'
        elif mark >= 75:
            return 'B+'
        elif mark >= 65:
            return 'B'
        elif mark >= 60:
            return 'B-'
        elif mark >= 55:
            return 'C'
        elif mark >= 50:
            return 'C-'
        elif mark >= 40:
            return 'D'
        elif mark >= 30:
            return 'D-'
        else:
            return 'E'
    else:
        # Theory grading
        if mark >= 85:
            return 'A+'
        elif mark >= 80:
            return 'A'
        elif mark >= 70:
            return 'B'
        elif mark >= 60:
            return 'B-'
        elif mark >= 50:
            return 'C'
        elif mark >= 40:
            return 'C-'
        elif mark >= 30:
            return 'D'
        else:
            return 'E'
