from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.db import transaction, IntegrityError
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django_countries import countries
from .models import Candidate, CandidateEnrollment, EnrollmentModule, EnrollmentPaper, CandidateActivity
from .serializers import (
    CandidateListSerializer,
    CandidateDetailSerializer,
    CandidateCreateUpdateSerializer,
    CandidateActivitySerializer,
    CandidateEnrollmentSerializer,
    EnrollmentListSerializer,
    EnrollCandidateSerializer,
    BulkEnrollSerializer
)
from occupations.models import OccupationLevel, OccupationModule, OccupationPaper
from assessment_series.models import AssessmentSeries


class CandidateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing candidates
    """
    permission_classes = [AllowAny]  # TODO: Change to IsAuthenticated after auth is set up
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['registration_number', 'full_name', 'contact', 'refugee_number']
    ordering_fields = ['created_at', 'full_name', 'registration_number']
    ordering = ['-created_at']
    
    filterset_fields = {
        'registration_category': ['exact'],
        'assessment_center': ['exact'],
        'occupation': ['exact'],
        'has_disability': ['exact'],
        'is_refugee': ['exact'],
        'verification_status': ['exact'],
        'status': ['exact'],
        'gender': ['exact'],
        'entry_year': ['exact'],
        'intake': ['exact'],
    }
    
    def get_queryset(self):
        """Optimize queryset with select_related and prefetch_related"""
        queryset = Candidate.objects.select_related(
            'district',
            'village',
            'nature_of_disability',
            'assessment_center',
            'assessment_center_branch',
            'occupation',
            'occupation__sector',
            'created_by',
            'updated_by',
            'verified_by'
        ).prefetch_related(
            'enrollments'
        ).all()
        
        # Custom filters for is_enrolled and has_marks
        is_enrolled = self.request.query_params.get('is_enrolled')
        if is_enrolled is not None:
            if is_enrolled.lower() == 'yes':
                queryset = queryset.filter(enrollments__is_active=True).distinct()
            elif is_enrolled.lower() == 'no':
                queryset = queryset.exclude(enrollments__is_active=True)
        
        has_marks = self.request.query_params.get('has_marks')
        if has_marks is not None:
            if has_marks.lower() == 'yes':
                queryset = queryset.filter(
                    Q(formal_results__isnull=False) | Q(modular_results__isnull=False)
                ).distinct()
            elif has_marks.lower() == 'no':
                queryset = queryset.exclude(
                    Q(formal_results__isnull=False) | Q(modular_results__isnull=False)
                )
        
        # Filter by center for center representatives
        if self.request.user.is_authenticated and self.request.user.user_type == 'center_representative':
            try:
                center_rep = self.request.user.center_rep_profile
                if center_rep and center_rep.assessment_center:
                    queryset = queryset.filter(assessment_center=center_rep.assessment_center)
                    # If center rep is assigned to a specific branch, filter by that branch too
                    if center_rep.assessment_center_branch:
                        queryset = queryset.filter(assessment_center_branch=center_rep.assessment_center_branch)
                else:
                    # No center assigned - return empty queryset for safety
                    queryset = queryset.none()
            except Exception:
                # No center_rep_profile found - return empty queryset for safety
                queryset = queryset.none()
        
        return queryset

    def _log_activity(self, candidate, action, description='', details=None):
        actor = self.request.user if getattr(self.request, 'user', None) and self.request.user.is_authenticated else None
        CandidateActivity.objects.create(
            candidate=candidate,
            actor=actor,
            action=action,
            description=description or '',
            details=details,
        )

    @action(detail=False, methods=['get'])
    def nationalities(self, request):
        """Return list of countries with East Africa prioritized"""
        # East African countries (using ISO codes)
        east_africa_codes = ['UG', 'KE', 'TZ', 'RW', 'BI', 'SS']
        
        # Get East African countries first (in specific order)
        east_africa = [
            {'value': code, 'label': countries.name(code)} 
            for code in east_africa_codes
        ]
        
        # Get remaining countries (alphabetically by name)
        remaining_countries = [
            {'value': code, 'label': name}
            for code, name in countries
            if code not in east_africa_codes
        ]
        # Sort remaining by label (country name)
        remaining_countries.sort(key=lambda x: x['label'])
        
        # Combine: East Africa first, then rest
        ordered = east_africa + remaining_countries
        
        # Optionally add 'Other' at the end
        ordered.append({'value': 'OTHER', 'label': 'Other'})
        
        return Response(ordered)
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return CandidateListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return CandidateCreateUpdateSerializer
        return CandidateDetailSerializer
    
    def perform_create(self, serializer):
        """Set created_by when creating a candidate"""
        staff = None
        try:
            staff = self.request.user.staff
        except Exception:
            pass
        if self.request.user.is_authenticated and self.request.user.user_type == 'center_representative':
            try:
                center_rep = self.request.user.center_rep_profile
                if center_rep and center_rep.assessment_center:
                    candidate = serializer.save(
                        created_by=staff,
                        updated_by=staff,
                        assessment_center=center_rep.assessment_center,
                        assessment_center_branch=center_rep.assessment_center_branch,
                    )
                    self._log_activity(candidate, 'candidate_created', 'Candidate created')
                    return
            except Exception:
                pass
        candidate = serializer.save(created_by=staff, updated_by=staff)
        self._log_activity(candidate, 'candidate_created', 'Candidate created')
    
    def perform_update(self, serializer):
        """Set updated_by when updating a candidate"""
        staff = None
        try:
            staff = self.request.user.staff
        except Exception:
            pass
        if self.request.user.is_authenticated and self.request.user.user_type == 'center_representative':
            try:
                center_rep = self.request.user.center_rep_profile
                if center_rep and center_rep.assessment_center:
                    candidate = serializer.save(
                        updated_by=staff,
                        assessment_center=center_rep.assessment_center,
                        assessment_center_branch=center_rep.assessment_center_branch,
                    )
                    self._log_activity(candidate, 'candidate_updated', 'Candidate updated')
                    return
            except Exception:
                pass
        candidate = serializer.save(updated_by=staff)
        self._log_activity(candidate, 'candidate_updated', 'Candidate updated')

    @action(detail=True, methods=['get'])
    def activity(self, request, pk=None):
        candidate = self.get_object()
        qs = candidate.activities.select_related('actor').all()
        serializer = CandidateActivitySerializer(qs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify a candidate"""
        candidate = self.get_object()
        staff = None
        if hasattr(request.user, 'staff'):
            staff = request.user.staff
        
        candidate.verification_status = 'verified'
        candidate.verified_by = staff
        candidate.verification_date = timezone.now()
        candidate.save()

        self._log_activity(candidate, 'candidate_verified', 'Candidate verified')
        
        serializer = self.get_serializer(candidate)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def decline(self, request, pk=None):
        """Decline a candidate"""
        candidate = self.get_object()
        staff = None
        if hasattr(request.user, 'staff'):
            staff = request.user.staff
        
        reason = request.data.get('reason', '')
        candidate.verification_status = 'declined'
        candidate.verified_by = staff
        candidate.decline_reason = reason
        candidate.verification_date = timezone.now()
        candidate.save()

        self._log_activity(candidate, 'candidate_declined', 'Candidate declined', details={'reason': reason})
        
        serializer = self.get_serializer(candidate)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def clear_payment(self, request, pk=None):
        """Mark payment as cleared for a candidate"""
        candidate = self.get_object()
        staff = None
        if hasattr(request.user, 'staff'):
            staff = request.user.staff
        
        candidate.payment_cleared = True
        candidate.payment_cleared_date = timezone.now().date()
        candidate.payment_cleared_by = staff
        candidate.payment_amount_cleared = request.data.get('amount', 0)
        candidate.payment_center_series_ref = request.data.get('reference', '')
        candidate.save()

        self._log_activity(candidate, 'payment_cleared', 'Payment cleared', details={
            'amount': request.data.get('amount', 0),
            'reference': request.data.get('reference', ''),
        })
        
        serializer = self.get_serializer(candidate)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def upload_document(self, request, pk=None):
        """Upload identification or qualification document"""
        candidate = self.get_object()
        
        if 'identification_document' in request.FILES:
            candidate.identification_document = request.FILES['identification_document']
        elif 'qualification_document' in request.FILES:
            candidate.qualification_document = request.FILES['qualification_document']
        else:
            return Response(
                {'error': 'No document file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        candidate.save()
        serializer = self.get_serializer(candidate)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get candidate statistics"""
        queryset = self.filter_queryset(self.get_queryset())
        
        stats = {
            'total': queryset.count(),
            'by_category': {
                'modular': queryset.filter(registration_category='modular').count(),
                'formal': queryset.filter(registration_category='formal').count(),
                'workers_pas': queryset.filter(registration_category='workers_pas').count(),
            },
            'by_status': {
                'verified': queryset.filter(verification_status='verified').count(),
                'pending': queryset.filter(verification_status='pending_verification').count(),
                'declined': queryset.filter(verification_status='declined').count(),
            },
            'by_gender': {
                'male': queryset.filter(gender='male').count(),
                'female': queryset.filter(gender='female').count(),
            },
            'refugees': queryset.filter(is_refugee=True).count(),
            'with_disability': queryset.filter(has_disability=True).count(),
        }
        
        return Response(stats)
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """Export candidates to Excel - optimized for large datasets"""
        candidate_ids = request.data.get('ids', [])
        export_all = request.data.get('export_all', False)
        
        # Get queryset based on filters or IDs - use values() for speed
        if export_all:
            # Apply filters from request data
            queryset = self.get_queryset()
            
            # Apply filters manually from POST data
            if request.data.get('registration_category'):
                queryset = queryset.filter(registration_category=request.data['registration_category'])
            if request.data.get('assessment_center'):
                queryset = queryset.filter(assessment_center_id=request.data['assessment_center'])
            if request.data.get('occupation'):
                queryset = queryset.filter(occupation_id=request.data['occupation'])
            if request.data.get('has_disability'):
                queryset = queryset.filter(has_disability=request.data['has_disability'] == 'true')
            if request.data.get('is_refugee'):
                queryset = queryset.filter(is_refugee=request.data['is_refugee'] == 'true')
            if request.data.get('verification_status'):
                queryset = queryset.filter(verification_status=request.data['verification_status'])
            if request.data.get('search'):
                search = request.data['search']
                queryset = queryset.filter(
                    Q(registration_number__icontains=search) |
                    Q(full_name__icontains=search) |
                    Q(contact__icontains=search)
                )

            # Add missing filters
            if request.data.get('entry_year'):
                queryset = queryset.filter(entry_year=request.data['entry_year'])
            if request.data.get('intake'):
                queryset = queryset.filter(intake=request.data['intake'])
            if request.data.get('sector'):
                queryset = queryset.filter(occupation__sector_id=request.data['sector'])
            
            # Custom filters for is_enrolled and has_marks (adapted from get_queryset)
            is_enrolled = request.data.get('is_enrolled')
            if is_enrolled is not None and str(is_enrolled): # Handle potential "undefined" or empty string
                if str(is_enrolled).lower() == 'yes':
                    queryset = queryset.filter(enrollments__is_active=True).distinct()
                elif str(is_enrolled).lower() == 'no':
                    queryset = queryset.exclude(enrollments__is_active=True)
            
            has_marks = request.data.get('has_marks')
            if has_marks is not None and str(has_marks):
                if str(has_marks).lower() == 'yes':
                    queryset = queryset.filter(
                        Q(formal_results__isnull=False) | Q(modular_results__isnull=False)
                    ).distinct()
                elif str(has_marks).lower() == 'no':
                    queryset = queryset.exclude(
                        Q(formal_results__isnull=False) | Q(modular_results__isnull=False)
                    )
        elif candidate_ids:
            queryset = self.get_queryset().filter(id__in=candidate_ids)
        else:
            return Response({'error': 'No candidates selected'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Use values for faster data retrieval (avoid model instantiation)
        candidates = queryset.values(
            'registration_number', 'full_name', 'date_of_birth', 'gender',
            'nationality', 'contact', 'has_disability', 'is_refugee',
            'assessment_center__center_name', 'registration_category',
            'occupation__occ_name', 'occupation__sector__name', 'district__name'
        )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"
        
        # Define headers and column widths
        headers = [
            ('Reg No', 20), ('Full Name', 25), ('Center', 30), ('Category', 12), 
            ('Occupation', 20), ('Sector', 15), ('Disability', 10), ('Refugee', 10), 
            ('Nationality', 12), ('Age', 6), ('District', 15), ('Gender', 8), ('Contact', 15)
        ]
        
        # Style for headers only
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        # Write headers and set column widths
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = width
        
        # Category display mapping
        category_map = {'modular': 'Modular', 'formal': 'Formal', 'workers_pas': "Worker's PAS"}
        gender_map = {'male': 'Male', 'female': 'Female', 'other': 'Other'}
        
        # Calculate age helper
        today = date.today()
        def calc_age(dob):
            if not dob:
                return ''
            return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        
        # Write data rows (no styling for speed)
        for row_num, c in enumerate(candidates, 2):
            ws.cell(row=row_num, column=1, value=c['registration_number'] or '')
            ws.cell(row=row_num, column=2, value=c['full_name'] or '')
            ws.cell(row=row_num, column=3, value=c['assessment_center__center_name'] or '')
            ws.cell(row=row_num, column=4, value=category_map.get(c['registration_category'], ''))
            ws.cell(row=row_num, column=5, value=c['occupation__occ_name'] or '')
            ws.cell(row=row_num, column=6, value=c['occupation__sector__name'] or '')
            ws.cell(row=row_num, column=7, value='Yes' if c['has_disability'] else 'No')
            ws.cell(row=row_num, column=8, value='Yes' if c['is_refugee'] else 'No')
            ws.cell(row=row_num, column=9, value=c['nationality'] or 'Uganda')
            ws.cell(row=row_num, column=10, value=calc_age(c['date_of_birth']))
            ws.cell(row=row_num, column=11, value=c['district__name'] or '')
            ws.cell(row=row_num, column=12, value=gender_map.get(c['gender'], ''))
            ws.cell(row=row_num, column=13, value=c['contact'] or '')
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=candidates_export_{date.today().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        
        return response
    
    @action(detail=True, methods=['get'])
    def enrollments(self, request, pk=None):
        """Get all enrollments for a candidate"""
        candidate = self.get_object()
        enrollments = candidate.enrollments.select_related(
            'assessment_series',
            'occupation_level',
            'occupation_level__occupation'
        ).prefetch_related('modules', 'papers').all()
        serializer = CandidateEnrollmentSerializer(enrollments, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def enroll(self, request, pk=None):
        """Enroll a candidate in an assessment series"""
        candidate = self.get_object()
        
        # Validate input
        serializer = EnrollCandidateSerializer(
            data=request.data,
            context={'candidate': candidate}
        )
        serializer.is_valid(raise_exception=True)
        
        validated_data = serializer.validated_data
        assessment_series = validated_data['assessment_series_obj']
        occupation_level = validated_data['occupation_level_obj']
        modules = validated_data.get('modules', [])
        papers = validated_data.get('papers', [])
        
        # Calculate billing (check if series has don't charge enabled)
        total_amount = Decimal('0.00')
        reg_category = candidate.registration_category
        
        # Check if assessment series has "don't charge" enabled
        if assessment_series.dont_charge:
            total_amount = Decimal('0.00')
        elif reg_category == 'formal':
            # Formal: use formal_fee from the level
            total_amount = occupation_level.formal_fee
        
        elif reg_category == 'modular':
            # Modular: use modular fee based on number of modules
            if len(modules) == 1:
                total_amount = occupation_level.modular_fee_single_module
            elif len(modules) == 2:
                total_amount = occupation_level.modular_fee_double_module
        
        elif reg_category == 'workers_pas':
            # Workers PAS: fee per paper (75k per paper)
            # Get the per-paper fee from any level (they should all be the same)
            any_level = candidate.occupation.levels.first()
            if any_level:
                per_paper_fee = any_level.workers_pas_per_module_fee  # Using this as per-paper fee
                total_amount = per_paper_fee * len(papers)
        
        # Create enrollment with transaction
        try:
            with transaction.atomic():
                # Check if enrollment already exists
                # For formal and modular, check by candidate + assessment_series (no duplicate enrollments allowed)
                # For workers_pas, also check by candidate + assessment_series
                existing = CandidateEnrollment.objects.filter(
                    candidate=candidate,
                    assessment_series=assessment_series,
                    is_active=True
                ).first()
                
                if existing:
                    level_info = f" at {existing.occupation_level.level_name}" if existing.occupation_level else ""
                    return Response(
                        {'error': f'Candidate is already enrolled in {assessment_series.name}{level_info}. Please de-enroll first.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Create enrollment
                enrollment = CandidateEnrollment.objects.create(
                    candidate=candidate,
                    assessment_series=assessment_series,
                    occupation_level=occupation_level,  # Will be None for Workers PAS
                    total_amount=total_amount
                )
                
                # For formal candidates, auto-enroll in all modules/papers for the level
                if reg_category == 'formal':
                    if occupation_level.structure_type == 'modules':
                        # Enroll in all modules for this level
                        all_modules = occupation_level.modules.all()
                        for module in all_modules:
                            EnrollmentModule.objects.create(
                                enrollment=enrollment,
                                module=module
                            )
                    elif occupation_level.structure_type == 'papers':
                        # Enroll in all papers for this level
                        all_papers = occupation_level.papers.all()
                        for paper in all_papers:
                            EnrollmentPaper.objects.create(
                                enrollment=enrollment,
                                paper=paper
                            )
                else:
                    # For modular/workers_pas, use the selected modules/papers
                    # Add modules
                    if modules:
                        for module_id in modules:
                            module = OccupationModule.objects.get(id=module_id)
                            EnrollmentModule.objects.create(
                                enrollment=enrollment,
                                module=module
                            )
                    
                    # Add papers
                    if papers:
                        for paper_id in papers:
                            paper = OccupationPaper.objects.get(id=paper_id)
                            EnrollmentPaper.objects.create(
                                enrollment=enrollment,
                                paper=paper
                            )
                
                # Return created enrollment
                enrollment_serializer = CandidateEnrollmentSerializer(enrollment)

                self._log_activity(candidate, 'candidate_enrolled', 'Candidate enrolled', details={
                    'assessment_series_id': assessment_series.id,
                    'assessment_series_name': assessment_series.name,
                    'enrollment_id': enrollment.id,
                    'registration_category': reg_category,
                    'total_amount': str(total_amount),
                })
                return Response(enrollment_serializer.data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-enroll')
    def bulk_enroll(self, request):
        """Bulk enroll multiple candidates in the same assessment series and level"""
        
        serializer = BulkEnrollSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        candidate_ids = serializer.validated_data['candidate_ids']
        assessment_series = serializer.validated_data['assessment_series_obj']
        occupation_level = serializer.validated_data.get('occupation_level_obj')
        
        # Validate all candidates have same registration category and occupation
        candidates = Candidate.objects.filter(id__in=candidate_ids).select_related('occupation')
        
        if len(candidates) != len(candidate_ids):
            return Response(
                {'error': 'Some candidates not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check all candidates have same registration category and occupation
        reg_categories = set(c.registration_category for c in candidates)
        occupations = set(c.occupation.id for c in candidates)
        
        if len(reg_categories) > 1:
            return Response(
                {'error': 'All candidates must have the same registration category'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(occupations) > 1:
            return Response(
                {'error': 'All candidates must have the same occupation'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reg_category = reg_categories.pop()
        occupation_id = occupations.pop()
        
        # Validate modules/papers for modular/workers_pas
        modules = serializer.validated_data.get('modules', [])
        papers = serializer.validated_data.get('papers', [])
        
        if reg_category == 'modular' and not modules:
            return Response(
                {'error': 'Modules are required for modular enrollment'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if reg_category == 'workers_pas' and not papers:
            return Response(
                {'error': 'Papers are required for workers_pas enrollment'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if level belongs to the occupation (only for formal/modular)
        if occupation_level and occupation_level.occupation.id != occupation_id:
            return Response(
                {'error': 'Level does not belong to the candidates occupation'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # For formal/modular, level is required
        if reg_category in ['formal', 'modular'] and not occupation_level:
            return Response(
                {'error': 'Level is required for formal/modular enrollment'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Bulk enroll candidates
        enrolled_count = 0
        failed_enrollments = []
        
        with transaction.atomic():
            for candidate in candidates:
                try:
                    # Check if already enrolled in this assessment series
                    existing = CandidateEnrollment.objects.filter(
                        candidate=candidate,
                        assessment_series=assessment_series,
                        is_active=True
                    ).first()
                    
                    if existing:
                        level_info = f" at {existing.occupation_level.level_name}" if existing.occupation_level else ""
                        failed_enrollments.append({
                            'candidate_id': candidate.id,
                            'name': candidate.full_name,
                            'reason': f'Already enrolled in {assessment_series.name}{level_info}. De-enroll first.'
                        })
                        continue
                    
                    # Calculate billing (check if series has don't charge enabled)
                    total_amount = Decimal('0.00')
                    
                    if assessment_series.dont_charge:
                        total_amount = Decimal('0.00')
                    elif reg_category == 'formal':
                        total_amount = occupation_level.formal_fee
                    elif reg_category == 'modular':
                        if len(modules) == 1:
                            total_amount = occupation_level.modular_fee_single_module
                        elif len(modules) == 2:
                            total_amount = occupation_level.modular_fee_double_module
                    elif reg_category == 'workers_pas':
                        any_level = candidate.occupation.levels.first()
                        if any_level:
                            per_paper_fee = any_level.workers_pas_per_module_fee
                            total_amount = per_paper_fee * len(papers)
                    
                    # Create enrollment
                    enrollment = CandidateEnrollment.objects.create(
                        candidate=candidate,
                        assessment_series=assessment_series,
                        occupation_level=occupation_level,
                        total_amount=total_amount
                    )
                    
                    # Add modules/papers for modular/workers_pas
                    if reg_category == 'modular' and modules:
                        for module in modules:
                            EnrollmentModule.objects.create(
                                enrollment=enrollment,
                                module=module
                            )
                    elif reg_category == 'workers_pas' and papers:
                        for paper in papers:
                            EnrollmentPaper.objects.create(
                                enrollment=enrollment,
                                paper=paper
                            )
                    
                    enrolled_count += 1
                    
                except Exception as e:
                    failed_enrollments.append({
                        'candidate_id': candidate.id,
                        'name': candidate.full_name,
                        'reason': str(e)
                    })
        
        # Return results
        if occupation_level:
            message = f'Successfully enrolled {enrolled_count} candidates in {occupation_level.level_name} {occupation_level.occupation.occ_name}'
            level_info = {
                'id': occupation_level.id,
                'name': occupation_level.level_name,
                'occupation': occupation_level.occupation.occ_name
            }
        else:
            # For workers_pas without occupation_level
            occupation = candidates[0].occupation
            message = f'Successfully enrolled {enrolled_count} candidates in {occupation.occ_name}'
            level_info = None
        
        response_data = {
            'message': message,
            'enrolled_count': enrolled_count,
            'total_candidates': len(candidate_ids),
        }
        
        if level_info:
            response_data['level'] = level_info
        
        if failed_enrollments:
            response_data['failed_enrollments'] = failed_enrollments
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'], url_path='enrollment-options')
    def enrollment_options(self, request, pk=None):
        """Get enrollment options for a candidate based on their registration category"""
        candidate = self.get_object()
        occupation = candidate.occupation
        
        if not occupation:
            return Response(
                {'error': 'Candidate has no occupation assigned'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reg_category = candidate.registration_category
        
        # Get assessment series
        # For center representatives, only show the current series
        if request.user.is_authenticated and request.user.user_type == 'center_representative':
            assessment_series = AssessmentSeries.objects.filter(is_active=True, is_current=True).values('id', 'name', 'start_date', 'end_date')
        else:
            assessment_series = AssessmentSeries.objects.filter(is_active=True).values('id', 'name', 'start_date', 'end_date')
        
        response_data = {
            'assessment_series': list(assessment_series),
            'registration_category': reg_category,
            'occupation': {
                'id': occupation.id,
                'occ_code': occupation.occ_code,
                'occ_name': occupation.occ_name,
            }
        }
        
        if reg_category == 'formal':
            # Formal: show all levels
            levels = occupation.levels.filter(is_active=True).values(
                'id', 'level_name', 'formal_fee'
            )
            response_data['levels'] = list(levels)
        
        elif reg_category == 'modular':
            # Modular: show level 1 with its modules
            level_1 = occupation.levels.filter(level_name__icontains='level 1', is_active=True).first()
            if level_1:
                modules = level_1.modules.filter(is_active=True).values(
                    'id', 'module_code', 'module_name'
                )
                response_data['level'] = {
                    'id': level_1.id,
                    'level_name': level_1.level_name,
                    'modular_fee_single_module': str(level_1.modular_fee_single_module),
                    'modular_fee_double_module': str(level_1.modular_fee_double_module),
                }
                response_data['modules'] = list(modules)
                response_data['module_limit'] = {'min': 1, 'max': 2}
        
        elif reg_category == 'workers_pas':
            # Workers PAS: show all levels with their modules and nested papers
            levels_data = []
            for level in occupation.levels.filter(is_active=True):
                modules_data = []
                for module in level.modules.filter(is_active=True):
                    # Get papers that belong to this module
                    papers = module.papers.filter(is_active=True).values(
                        'id', 'paper_code', 'paper_name', 'paper_type'
                    )
                    modules_data.append({
                        'id': module.id,
                        'module_code': module.module_code,
                        'module_name': module.module_name,
                        'papers': list(papers),
                    })
                
                levels_data.append({
                    'id': level.id,
                    'level_name': level.level_name,
                    'workers_pas_base_fee': str(level.workers_pas_base_fee),
                    'workers_pas_per_module_fee': str(level.workers_pas_per_module_fee),
                    'modules': modules_data,
                })
            response_data['levels'] = levels_data
        
        return Response(response_data)
    
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """
        Submit a draft candidate and generate registration number
        Format: UVT001/X/25/M/HD/M/001
        - UVT001: Center number
        - X: Nationality (U for Uganda, X for others)
        - 25: Entry year (last 2 digits)
        - M: Assessment intake (M, J, S, D)
        - HD: Occupation code
        - M: Registration category (M=Modular, F=Formal, I=Workers PAS)
        - 001: Unique sequence per center+occupation
        """
        candidate = self.get_object()
        
        if candidate.is_submitted:
            return Response(
                {'error': 'Candidate has already been submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate required fields
        if not candidate.assessment_center:
            return Response(
                {'error': 'Assessment center is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not candidate.occupation:
            return Response(
                {'error': 'Occupation is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not candidate.entry_year:
            return Response(
                {'error': 'Entry year is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not candidate.intake:
            return Response(
                {'error': 'Intake is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not candidate.registration_category:
            return Response(
                {'error': 'Registration category is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build registration number components
        # 1. Center number (e.g., UVT001)
        center_number = candidate.assessment_center.center_number
        
        # 2. Nationality code (U for Uganda, X for others)
        nationality_code = candidate.get_nationality_code()
        
        # 3. Entry year (last 2 digits)
        year_code = str(candidate.entry_year)[-2:]
        
        # 4. Intake (M, J, S, D)
        intake_code = candidate.intake
        
        # 5. Occupation code
        occ_code = candidate.occupation.occ_code
        
        # 6. Registration category code
        reg_cat_code = candidate.get_registration_category_code()
        
        # 7. Sequence number (unique per center + occupation)
        # Find the MAX sequence from all existing registration numbers for this center+occupation
        existing_candidates = Candidate.objects.filter(
            assessment_center=candidate.assessment_center,
            occupation=candidate.occupation,
            is_submitted=True,
            registration_number__isnull=False
        ).exclude(pk=candidate.pk).values_list('registration_number', flat=True)
        
        max_seq = 0
        for reg_num in existing_candidates:
            if reg_num:
                parts = reg_num.split('/')
                try:
                    seq = int(parts[-1])
                    if seq > max_seq:
                        max_seq = seq
                except (ValueError, IndexError):
                    pass
        
        new_seq = max_seq + 1
        
        # Generate registration number with retry for race conditions
        max_retries = 5
        for attempt in range(max_retries):
            registration_number = f'{center_number}/{nationality_code}/{year_code}/{intake_code}/{occ_code}/{reg_cat_code}/{new_seq:04d}'
            candidate.registration_number = registration_number
            candidate.is_submitted = True
            try:
                candidate.save()
                break
            except IntegrityError:
                if attempt < max_retries - 1:
                    # Collision detected, increment sequence and retry
                    new_seq += 1
                    continue
                else:
                    raise
        
        # Generate payment code after saving (so candidate.pk is available)
        payment_code = candidate.generate_payment_code()
        if payment_code:
            candidate.payment_code = payment_code
            candidate.save()

        self._log_activity(candidate, 'candidate_submitted', 'Candidate submitted', details={
            'registration_number': registration_number,
            'payment_code': payment_code,
        })
        
        return Response(
            {
                'message': 'Candidate submitted successfully',
                'registration_number': registration_number,
                'payment_code': payment_code
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_photo(self, request, pk=None):
        """Upload passport photo"""
        candidate = self.get_object()
        
        photo_file = request.FILES.get('photo') or request.FILES.get('passport_photo')
        if not photo_file:
            return Response(
                {'error': 'No photo file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        candidate.passport_photo = photo_file
        candidate.save()

        self._log_activity(candidate, 'photo_uploaded', 'Photo uploaded')
        
        return Response(
            {
                'message': 'Photo uploaded successfully',
                'photo_url': candidate.passport_photo.url if candidate.passport_photo else None
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['delete'])
    def delete_photo(self, request, pk=None):
        """Delete passport photo"""
        candidate = self.get_object()

        if not candidate.passport_photo:
            return Response(
                {'error': 'No photo to delete'},
                status=status.HTTP_400_BAD_REQUEST
            )

        candidate.passport_photo.delete(save=False)
        candidate.passport_photo = None
        candidate.save()

        self._log_activity(candidate, 'photo_deleted', 'Passport photo deleted')

        return Response(
            {'message': 'Photo deleted successfully'},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_document(self, request, pk=None):
        """Upload identification or qualification document"""
        candidate = self.get_object()
        
        document_type = request.data.get('document_type')
        
        if document_type not in ['identification', 'qualification']:
            return Response(
                {'error': 'Invalid document type. Must be "identification" or "qualification"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'document' not in request.FILES:
            return Response(
                {'error': 'No document file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if document_type == 'identification':
            candidate.identification_document = request.FILES['document']
        else:
            candidate.qualification_document = request.FILES['document']
        
        candidate.save()

        self._log_activity(candidate, 'document_uploaded', f'{document_type.capitalize()} document uploaded', details={
            'document_type': document_type,
        })
        
        return Response(
            {
                'message': f'{document_type.capitalize()} document uploaded successfully',
                'document_url': (
                    candidate.identification_document.url if document_type == 'identification'
                    else candidate.qualification_document.url
                ) if (candidate.identification_document or candidate.qualification_document) else None
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def generate_payment_code(self, request, pk=None):
        """Generate payment code for a candidate who has a registration number"""
        candidate = self.get_object()
        
        # Check if candidate is submitted and has registration number
        if not candidate.is_submitted:
            return Response(
                {'error': 'Candidate must be submitted first'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not candidate.registration_number:
            return Response(
                {'error': 'Candidate must have a registration number'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if payment code already exists
        if candidate.payment_code:
            return Response(
                {
                    'message': 'Payment code already exists',
                    'payment_code': candidate.payment_code
                },
                status=status.HTTP_200_OK
            )
        
        # Validate required fields
        if not all([candidate.assessment_center, candidate.entry_year, candidate.pk]):
            return Response(
                {'error': 'Missing required fields (assessment center, entry year, or candidate ID)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate payment code
        payment_code = candidate.generate_payment_code()
        
        if not payment_code:
            return Response(
                {'error': 'Failed to generate payment code'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Save payment code
        candidate.payment_code = payment_code
        candidate.save()

        self._log_activity(candidate, 'payment_code_generated', 'Payment code generated', details={'payment_code': payment_code})
        
        return Response(
            {
                'message': 'Payment code generated successfully',
                'payment_code': payment_code
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def mark_payment_cleared(self, request, pk=None):
        """Mark candidate payment as cleared (for manual payment confirmation)"""
        candidate = self.get_object()
        
        # Check if candidate has enrollments
        if not candidate.enrollments.exists():
            return Response(
                {'error': 'Candidate has no enrollments to bill'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Calculate total billed amount
        from django.db.models import Sum
        from decimal import Decimal
        total_billed = candidate.enrollments.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        # Mark payment as cleared
        candidate.payment_cleared = True
        candidate.payment_amount_cleared = total_billed
        candidate.payment_cleared_date = timezone.now()
        candidate.payment_cleared_by = request.user.get_full_name() or request.user.username
        
        # Set reference if not already set
        if not candidate.payment_center_series_ref:
            candidate.payment_center_series_ref = f'MANUAL-{timezone.now().strftime("%Y%m%d%H%M%S")}'
        
        candidate.save()

        self._log_activity(candidate, 'payment_marked_cleared', 'Payment marked as cleared', details={
            'amount_cleared': float(total_billed),
            'reference': candidate.payment_center_series_ref,
        })
        
        return Response(
            {
                'message': 'Payment marked as cleared successfully',
                'payment_cleared': True,
                'amount_cleared': float(total_billed),
                'cleared_by': candidate.payment_cleared_by,
                'cleared_date': candidate.payment_cleared_date
            },
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['get'])
    def results(self, request, pk=None):
        """Get candidate results based on registration category"""
        candidate = self.get_object()
        
        # Check if user is staff (can see all results) or center user (only released results)
        user = request.user
        is_staff_user = user.is_authenticated and user.user_type in ['staff', 'support_staff']
        
        # Get results based on registration category
        if candidate.registration_category == 'modular':
            from results.models import ModularResult
            results = ModularResult.objects.filter(candidate=candidate).select_related(
                'assessment_series', 'module', 'entered_by'
            )
            
            # For center users, only show results from released assessment series
            if not is_staff_user:
                results = results.filter(assessment_series__results_released=True)
            
            results_data = []
            for result in results:
                results_data.append({
                    'id': result.id,
                    'assessment_series_name': result.assessment_series.name,
                    'module_name': result.module.module_name,
                    'module_code': result.module.module_code,
                    'type': result.type,
                    'mark': float(result.mark) if result.mark is not None else None,
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                    'entered_by': result.entered_by.get_full_name() if result.entered_by else None,
                    'entered_at': result.entered_at,
                })
            
            return Response(results_data, status=status.HTTP_200_OK)
        
        elif candidate.registration_category == 'formal':
            from results.models import FormalResult
            results = FormalResult.objects.filter(candidate=candidate).select_related(
                'assessment_series', 'level', 'exam', 'paper', 'entered_by'
            ).order_by('level__level_name', 'type')
            
            # For center users, only show results from released assessment series
            if not is_staff_user:
                results = results.filter(assessment_series__results_released=True)
            
            results_data = []
            for result in results:
                exam_or_paper_name = ''
                if result.exam:
                    exam_or_paper_name = result.exam.module_name
                elif result.paper:
                    exam_or_paper_name = result.paper.paper_name
                
                results_data.append({
                    'id': result.id,
                    'assessment_series': {
                        'id': result.assessment_series.id,
                        'name': result.assessment_series.name,
                    },
                    'level': {
                        'id': result.level.id,
                        'name': result.level.level_name,
                        'structure_type': result.level.structure_type,
                    },
                    'exam_or_paper': exam_or_paper_name,
                    'is_exam': result.exam is not None,
                    'type': result.type,
                    'mark': float(result.mark) if result.mark is not None else None,
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                    'entered_by': result.entered_by.get_full_name() if result.entered_by else None,
                    'entered_at': result.entered_at,
                })
            
            return Response(results_data, status=status.HTTP_200_OK)
        
        elif candidate.registration_category == 'workers_pas':
            from results.models import WorkersPasResult
            results = WorkersPasResult.objects.filter(candidate=candidate).select_related(
                'assessment_series', 'level', 'module', 'paper', 'entered_by'
            ).order_by('assessment_series', 'level', 'module', 'paper')
            
            # For center users, only show results from released assessment series
            if not is_staff_user:
                results = results.filter(assessment_series__results_released=True)
            
            results_data = []
            for result in results:
                results_data.append({
                    'id': result.id,
                    'assessment_series_name': result.assessment_series.name,
                    'level_name': result.level.level_name,
                    'module_code': result.module.module_code,
                    'module_name': result.module.module_name,
                    'paper_code': result.paper.paper_code,
                    'paper_name': result.paper.paper_name,
                    'type': result.type,
                    'mark': float(result.mark) if result.mark is not None else None,
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                    'entered_by_name': result.entered_by.get_full_name() if result.entered_by else None,
                    'entered_at': result.entered_at,
                })
            
            return Response(results_data, status=status.HTTP_200_OK)
        
        # Placeholder for other categories
        return Response([], status=status.HTTP_200_OK)
    
    # Modular results endpoints moved to results/views.py
    # - add_results -> /api/results/modular/add/
    # - enrollment_modules -> /api/results/modular/enrollment-modules/
    # - update_results -> /api/results/modular/update/
    # - verified_results_pdf -> /api/results/modular/verified-pdf/


# Standalone view for de-enrollment
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_de_enroll_view(request):
    """Bulk de-enroll candidates by deleting their enrollments"""
    candidate_ids = request.data.get('candidate_ids', [])
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    success_count = 0
    failed = []
    skipped_with_marks = []
    
    # Get all enrollments for selected candidates
    enrollments = CandidateEnrollment.objects.filter(candidate_id__in=candidate_ids)
    
    for enrollment in enrollments:
        candidate = enrollment.candidate
        series = enrollment.assessment_series
        
        # Check if candidate has any results for this enrollment
        has_modular_results = ModularResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        has_formal_results = FormalResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        has_workers_pas_results = WorkersPasResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        if has_modular_results or has_formal_results or has_workers_pas_results:
            skipped_with_marks.append({
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has marks - cannot de-enroll'
            })
            continue
        
        # Check if candidate has marked/approved fees for this series
        from fees.models import CandidateFee
        has_locked_fees = CandidateFee.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series,
            verification_status__in=['marked', 'approved']
        ).exists()
        if has_locked_fees:
            skipped_with_marks.append({
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has fees marked/approved by accounts - cannot de-enroll'
            })
            continue
        
        try:
            enrollment.delete()
            actor = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
            CandidateActivity.objects.create(
                candidate=candidate,
                actor=actor,
                action='candidate_deenrolled',
                description='Candidate de-enrolled',
                details={
                    'enrollment_id': enrollment.id,
                    'assessment_series_id': series.id if series else None,
                    'assessment_series_name': series.name if series else None,
                },
            )
            success_count += 1
        except Exception as e:
            failed.append({
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reason': str(e)
            })
    
    return Response({
        'message': f'Successfully de-enrolled {success_count} candidate(s)',
        'success_count': success_count,
        'skipped_with_marks': skipped_with_marks,
        'failed': failed
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['DELETE'])
@permission_classes([AllowAny])
def delete_enrollment_view(request, enrollment_id):
    """De-enroll a candidate by deleting their enrollment"""
    try:
        enrollment = CandidateEnrollment.objects.get(id=enrollment_id)
        candidate = enrollment.candidate
        series = enrollment.assessment_series
        
        # Check if candidate has any results
        from results.models import ModularResult, FormalResult, WorkersPasResult
        
        has_modular_results = ModularResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        has_formal_results = FormalResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        has_workers_pas_results = WorkersPasResult.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series
        ).exists()
        
        if has_modular_results or has_formal_results or has_workers_pas_results:
            return Response(
                {'error': 'Can\'t de-enroll, candidate already has marks'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if candidate has marked/approved fees for this series
        from fees.models import CandidateFee
        has_locked_fees = CandidateFee.objects.filter(
            candidate=candidate,
            assessment_series=enrollment.assessment_series,
            verification_status__in=['marked', 'approved']
        ).exists()
        if has_locked_fees:
            return Response(
                {'error': 'Can\'t de-enroll, candidate has fees marked/approved by accounts'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        enrollment.delete()
        actor = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
        CandidateActivity.objects.create(
            candidate=candidate,
            actor=actor,
            action='enrollment_deleted',
            description='Enrollment deleted',
            details={
                'enrollment_id': enrollment_id,
                'assessment_series_id': series.id if series else None,
                'assessment_series_name': series.name if series else None,
            },
        )
        
        return Response(
            {'message': 'Enrollment deleted successfully'},
            status=status.HTTP_200_OK
        )
    except CandidateEnrollment.DoesNotExist:
        return Response(
            {'error': 'Enrollment not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def change_candidate_series(request, candidate_id):
    """Change assessment series for a candidate's enrollments and results"""
    new_series_id = request.data.get('new_series_id')
    
    if not new_series_id:
        return Response(
            {'error': 'New assessment series ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidate = Candidate.objects.get(id=candidate_id)
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from assessment_series.models import AssessmentSeries
    try:
        new_series = AssessmentSeries.objects.get(id=new_series_id)
    except AssessmentSeries.DoesNotExist:
        return Response(
            {'error': 'Assessment series not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    updated = {
        'enrollments': 0,
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
    }
    
    # Update all enrollments
    enrollments_updated = CandidateEnrollment.objects.filter(candidate=candidate).update(assessment_series=new_series)
    updated['enrollments'] = enrollments_updated
    
    # Update all modular results
    modular_updated = ModularResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
    updated['modular_results'] = modular_updated
    
    # Update all formal results
    formal_updated = FormalResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
    updated['formal_results'] = formal_updated
    
    # Update all workers PAS results
    workers_updated = WorkersPasResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
    updated['workers_pas_results'] = workers_updated
    
    return Response({
        'message': f'Successfully moved {candidate.full_name} to {new_series.name}',
        'candidate_id': candidate.id,
        'new_series': {
            'id': new_series.id,
            'name': new_series.name,
        },
        'updated': updated
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def change_candidate_center(request, candidate_id):
    """Change assessment center for a candidate, updating registration number and fees"""
    new_center_id = request.data.get('new_center_id')
    
    if not new_center_id:
        return Response(
            {'error': 'New assessment center ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidate = Candidate.objects.get(id=candidate_id)
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from assessment_centers.models import AssessmentCenter
    try:
        new_center = AssessmentCenter.objects.get(id=new_center_id)
    except AssessmentCenter.DoesNotExist:
        return Response(
            {'error': 'Assessment center not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    old_center = candidate.assessment_center
    old_registration_number = candidate.registration_number
    
    # Update candidate's assessment center
    candidate.assessment_center = new_center
    candidate.assessment_center_branch = None  # Reset branch since new center may have different branches
    
    # Generate new registration number
    new_registration_number = candidate.generate_registration_number()
    candidate.registration_number = new_registration_number
    
    # Regenerate payment code with new center
    new_payment_code = candidate.generate_payment_code()
    candidate.payment_code = new_payment_code
    
    candidate.save()
    
    # Update CenterFee totals if candidate has fees
    from fees.models import CandidateFee, CenterFee
    
    candidate_fees = CandidateFee.objects.filter(candidate=candidate)
    fees_moved = 0
    
    for candidate_fee in candidate_fees:
        if candidate_fee.total_amount > 0:
            fees_moved += 1
            
            # Decrease old center's fee totals
            if old_center:
                try:
                    old_center_fee = CenterFee.objects.get(
                        assessment_center=old_center,
                        assessment_series=candidate_fee.assessment_series
                    )
                    old_center_fee.total_candidates = max(0, old_center_fee.total_candidates - 1)
                    old_center_fee.total_amount = max(0, old_center_fee.total_amount - candidate_fee.total_amount)
                    old_center_fee.save()
                except CenterFee.DoesNotExist:
                    pass
            
            # Increase new center's fee totals
            new_center_fee, created = CenterFee.objects.get_or_create(
                assessment_center=new_center,
                assessment_series=candidate_fee.assessment_series,
                defaults={
                    'total_candidates': 0,
                    'total_amount': 0,
                    'amount_due': 0,
                }
            )
            new_center_fee.total_candidates += 1
            new_center_fee.total_amount += candidate_fee.total_amount
            new_center_fee.save()
    
    return Response({
        'message': f'Successfully moved {candidate.full_name} to {new_center.center_name}',
        'candidate_id': candidate.id,
        'old_center': {
            'id': old_center.id if old_center else None,
            'name': old_center.center_name if old_center else None,
        },
        'new_center': {
            'id': new_center.id,
            'name': new_center.center_name,
        },
        'old_registration_number': old_registration_number,
        'new_registration_number': new_registration_number,
        'fees_moved': fees_moved,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def change_candidate_occupation(request, candidate_id):
    """Change occupation for a candidate, updating registration number"""
    new_occupation_id = request.data.get('new_occupation_id')
    
    if not new_occupation_id:
        return Response(
            {'error': 'New occupation ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidate = Candidate.objects.get(id=candidate_id)
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from occupations.models import Occupation
    try:
        new_occupation = Occupation.objects.get(id=new_occupation_id)
    except Occupation.DoesNotExist:
        return Response(
            {'error': 'Occupation not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    old_occupation = candidate.occupation
    old_registration_number = candidate.registration_number
    
    # Check if candidate has any enrollments
    has_enrollments = CandidateEnrollment.objects.filter(candidate=candidate).exists()
    if has_enrollments:
        return Response(
            {'error': "Cannot change occupation. Candidate has existing enrollments. Clear enrollments first before changing occupation."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if candidate has any results
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    has_results = (
        ModularResult.objects.filter(candidate=candidate).exists() or
        FormalResult.objects.filter(candidate=candidate).exists() or
        WorkersPasResult.objects.filter(candidate=candidate).exists()
    )
    if has_results:
        return Response(
            {'error': "Cannot change occupation. Candidate has existing results. Clear results first before changing occupation."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate that new occupation supports candidate's registration category
    candidate_reg_cat = candidate.registration_category
    
    if candidate_reg_cat == 'workers_pas':
        # Workers PAS candidates can only move to workers_pas occupations
        if new_occupation.occ_category != 'workers_pas':
            return Response(
                {'error': f"Occupation doesn't support current registration category. {new_occupation.occ_name} is a {new_occupation.get_occ_category_display()} occupation, but candidate is registered as Worker's PAS."},
                status=status.HTTP_400_BAD_REQUEST
            )
    elif candidate_reg_cat == 'formal':
        # Formal candidates can only move to formal occupations
        if new_occupation.occ_category != 'formal':
            return Response(
                {'error': f"Occupation doesn't support current registration category. {new_occupation.occ_name} is a {new_occupation.get_occ_category_display()} occupation, but candidate is registered as Formal."},
                status=status.HTTP_400_BAD_REQUEST
            )
    elif candidate_reg_cat == 'modular':
        # Modular candidates can only move to formal occupations that support modular
        if new_occupation.occ_category != 'formal' or not new_occupation.has_modular:
            return Response(
                {'error': f"Occupation doesn't support current registration category. {new_occupation.occ_name} doesn't support Modular registration."},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    # Update candidate's occupation
    candidate.occupation = new_occupation
    
    # Generate new registration number
    new_registration_number = candidate.generate_registration_number()
    candidate.registration_number = new_registration_number
    
    candidate.save()
    
    return Response({
        'message': f'Successfully changed occupation for {candidate.full_name} to {new_occupation.occ_name}',
        'candidate_id': candidate.id,
        'old_occupation': {
            'id': old_occupation.id if old_occupation else None,
            'name': old_occupation.occ_name if old_occupation else None,
            'code': old_occupation.occ_code if old_occupation else None,
        },
        'new_occupation': {
            'id': new_occupation.id,
            'name': new_occupation.occ_name,
            'code': new_occupation.occ_code,
        },
        'old_registration_number': old_registration_number,
        'new_registration_number': new_registration_number,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def change_candidate_registration_category(request, candidate_id):
    """Change registration category for a candidate"""
    new_reg_category = request.data.get('new_registration_category')
    
    valid_categories = ['modular', 'formal', 'workers_pas']
    if not new_reg_category or new_reg_category not in valid_categories:
        return Response(
            {'error': f'Invalid registration category. Must be one of: {", ".join(valid_categories)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidate = Candidate.objects.get(id=candidate_id)
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    old_reg_category = candidate.registration_category
    
    # Check if trying to change to same category
    if old_reg_category == new_reg_category:
        return Response(
            {'error': 'Candidate is already registered in this category'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if candidate has any enrollments
    has_enrollments = CandidateEnrollment.objects.filter(candidate=candidate).exists()
    if has_enrollments:
        return Response(
            {'error': "Cannot change registration category. Candidate has existing enrollments. Clear enrollments first."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if candidate has any results
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    has_results = (
        ModularResult.objects.filter(candidate=candidate).exists() or
        FormalResult.objects.filter(candidate=candidate).exists() or
        WorkersPasResult.objects.filter(candidate=candidate).exists()
    )
    if has_results:
        return Response(
            {'error': "Cannot change registration category. Candidate has existing results. Clear results first."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if current occupation supports the new registration category
    occupation = candidate.occupation
    if not occupation:
        return Response(
            {'error': "Cannot change registration category. Candidate has no occupation assigned."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validation based on target registration category
    if new_reg_category == 'workers_pas':
        if occupation.occ_category != 'workers_pas':
            return Response(
                {'error': f"Cannot change to Worker's PAS. Current occupation '{occupation.occ_name}' is a {occupation.get_occ_category_display()} occupation and doesn't support Worker's PAS registration."},
                status=status.HTTP_400_BAD_REQUEST
            )
    elif new_reg_category == 'formal':
        if occupation.occ_category != 'formal':
            return Response(
                {'error': f"Cannot change to Formal. Current occupation '{occupation.occ_name}' is a {occupation.get_occ_category_display()} occupation and doesn't support Formal registration."},
                status=status.HTTP_400_BAD_REQUEST
            )
    elif new_reg_category == 'modular':
        if occupation.occ_category != 'formal':
            return Response(
                {'error': f"Cannot change to Modular. Current occupation '{occupation.occ_name}' is a {occupation.get_occ_category_display()} occupation and doesn't support Modular registration."},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not occupation.has_modular:
            return Response(
                {'error': f"Cannot change to Modular. Current occupation '{occupation.occ_name}' doesn't have Modular registration enabled."},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    old_registration_number = candidate.registration_number
    
    # Update registration category
    candidate.registration_category = new_reg_category
    
    # Generate new registration number (reg category code is part of it)
    new_registration_number = candidate.generate_registration_number()
    candidate.registration_number = new_registration_number
    
    candidate.save()
    
    category_display = {
        'modular': 'Modular',
        'formal': 'Formal',
        'workers_pas': "Worker's PAS"
    }
    
    return Response({
        'message': f'Successfully changed registration category for {candidate.full_name} to {category_display[new_reg_category]}',
        'candidate_id': candidate.id,
        'old_registration_category': old_reg_category,
        'new_registration_category': new_reg_category,
        'old_registration_number': old_registration_number,
        'new_registration_number': new_registration_number,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_change_candidate_occupation(request):
    """Bulk change occupation for multiple candidates"""
    candidate_ids = request.data.get('candidate_ids', [])
    new_occupation_id = request.data.get('new_occupation_id')
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not new_occupation_id:
        return Response(
            {'error': 'New occupation ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from occupations.models import Occupation
    try:
        new_occupation = Occupation.objects.get(id=new_occupation_id)
    except Occupation.DoesNotExist:
        return Response(
            {'error': 'Occupation not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    candidates = Candidate.objects.filter(id__in=candidate_ids)
    
    successful = []
    failed = []
    
    for candidate in candidates:
        # Check if candidate has any enrollments
        has_enrollments = CandidateEnrollment.objects.filter(candidate=candidate).exists()
        if has_enrollments:
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': 'Has existing enrollments'
            })
            continue
        
        # Check if candidate has any results
        has_results = (
            ModularResult.objects.filter(candidate=candidate).exists() or
            FormalResult.objects.filter(candidate=candidate).exists() or
            WorkersPasResult.objects.filter(candidate=candidate).exists()
        )
        if has_results:
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': 'Has existing results'
            })
            continue
        
        # Validate registration category compatibility
        candidate_reg_cat = candidate.registration_category
        
        if candidate_reg_cat == 'workers_pas' and new_occupation.occ_category != 'workers_pas':
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': f"Registration category mismatch (Worker's PAS → {new_occupation.get_occ_category_display()})"
            })
            continue
        elif candidate_reg_cat == 'formal' and new_occupation.occ_category != 'formal':
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': f"Registration category mismatch (Formal → {new_occupation.get_occ_category_display()})"
            })
            continue
        elif candidate_reg_cat == 'modular' and (new_occupation.occ_category != 'formal' or not new_occupation.has_modular):
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': "Occupation doesn't support Modular registration"
            })
            continue
        
        # All validations passed - update candidate
        old_registration_number = candidate.registration_number
        candidate.occupation = new_occupation
        candidate.registration_number = candidate.generate_registration_number()
        candidate.save()
        
        successful.append({
            'id': candidate.id,
            'name': candidate.full_name,
            'old_reg_no': old_registration_number,
            'new_reg_no': candidate.registration_number
        })
    
    return Response({
        'message': f'Changed occupation for {len(successful)} candidate(s) to {new_occupation.occ_name}',
        'successful': len(successful),
        'failed': len(failed),
        'successful_details': successful,
        'failed_details': failed,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_change_candidate_registration_category(request):
    """Bulk change registration category for multiple candidates"""
    candidate_ids = request.data.get('candidate_ids', [])
    new_reg_category = request.data.get('new_registration_category')
    
    valid_categories = ['modular', 'formal', 'workers_pas']
    if not new_reg_category or new_reg_category not in valid_categories:
        return Response(
            {'error': f'Invalid registration category. Must be one of: {", ".join(valid_categories)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    candidates = Candidate.objects.filter(id__in=candidate_ids)
    
    successful = []
    failed = []
    
    category_display = {
        'modular': 'Modular',
        'formal': 'Formal',
        'workers_pas': "Worker's PAS"
    }
    
    for candidate in candidates:
        # Check if already in target category
        if candidate.registration_category == new_reg_category:
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': f'Already registered as {category_display[new_reg_category]}'
            })
            continue
        
        # Check for enrollments
        if CandidateEnrollment.objects.filter(candidate=candidate).exists():
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': 'Has existing enrollments'
            })
            continue
        
        # Check for results
        has_results = (
            ModularResult.objects.filter(candidate=candidate).exists() or
            FormalResult.objects.filter(candidate=candidate).exists() or
            WorkersPasResult.objects.filter(candidate=candidate).exists()
        )
        if has_results:
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': 'Has existing results'
            })
            continue
        
        # Check occupation
        occupation = candidate.occupation
        if not occupation:
            failed.append({
                'id': candidate.id,
                'name': candidate.full_name,
                'reason': 'No occupation assigned'
            })
            continue
        
        # Validate occupation supports target category
        if new_reg_category == 'workers_pas':
            if occupation.occ_category != 'workers_pas':
                failed.append({
                    'id': candidate.id,
                    'name': candidate.full_name,
                    'reason': f"Occupation doesn't support Worker's PAS"
                })
                continue
        elif new_reg_category == 'formal':
            if occupation.occ_category != 'formal':
                failed.append({
                    'id': candidate.id,
                    'name': candidate.full_name,
                    'reason': "Occupation doesn't support Formal"
                })
                continue
        elif new_reg_category == 'modular':
            if occupation.occ_category != 'formal':
                failed.append({
                    'id': candidate.id,
                    'name': candidate.full_name,
                    'reason': "Occupation doesn't support Modular"
                })
                continue
            if not occupation.has_modular:
                failed.append({
                    'id': candidate.id,
                    'name': candidate.full_name,
                    'reason': "Occupation doesn't have Modular enabled"
                })
                continue
        
        old_registration_number = candidate.registration_number
        
        # Update registration category
        candidate.registration_category = new_reg_category
        candidate.registration_number = candidate.generate_registration_number()
        candidate.save()
        
        successful.append({
            'id': candidate.id,
            'name': candidate.full_name,
            'old_reg_no': old_registration_number,
            'new_reg_no': candidate.registration_number
        })
    
    return Response({
        'message': f'Changed registration category for {len(successful)} candidate(s) to {category_display[new_reg_category]}',
        'successful': len(successful),
        'failed': len(failed),
        'successful_details': successful,
        'failed_details': failed,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_change_candidate_series(request):
    """Bulk change assessment series for multiple candidates"""
    candidate_ids = request.data.get('candidate_ids', [])
    new_series_id = request.data.get('new_series_id')
    
    if not new_series_id:
        return Response(
            {'error': 'New assessment series ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from assessment_series.models import AssessmentSeries
    try:
        new_series = AssessmentSeries.objects.get(id=new_series_id)
    except AssessmentSeries.DoesNotExist:
        return Response(
            {'error': 'Assessment series not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    candidates = Candidate.objects.filter(id__in=candidate_ids)
    
    total_updated = {
        'candidates': 0,
        'enrollments': 0,
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
    }
    
    for candidate in candidates:
        # Update all enrollments
        enrollments_updated = CandidateEnrollment.objects.filter(candidate=candidate).update(assessment_series=new_series)
        total_updated['enrollments'] += enrollments_updated
        
        # Update all modular results
        modular_updated = ModularResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
        total_updated['modular_results'] += modular_updated
        
        # Update all formal results
        formal_updated = FormalResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
        total_updated['formal_results'] += formal_updated
        
        # Update all workers PAS results
        workers_updated = WorkersPasResult.objects.filter(candidate=candidate).update(assessment_series=new_series)
        total_updated['workers_pas_results'] += workers_updated
        
        total_updated['candidates'] += 1
    
    return Response({
        'message': f'Successfully moved {total_updated["candidates"]} candidate(s) to {new_series.name}',
        'new_series': {
            'id': new_series.id,
            'name': new_series.name,
        },
        'updated': total_updated
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_clear_candidate_data(request):
    """Bulk clear all results, enrollments, and fees for multiple candidates"""
    if request.user.is_authenticated and request.user.user_type == 'center_representative':
        return Response({'error': 'Permission denied. Center representatives cannot perform this action.'}, status=status.HTTP_403_FORBIDDEN)

    candidate_ids = request.data.get('candidate_ids', [])
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    from fees.models import CandidateFee
    
    total_cleared = {
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
        'enrollments': 0,
        'candidates_processed': 0,
    }
    skipped = []
    
    candidates = Candidate.objects.filter(id__in=candidate_ids)
    
    for candidate in candidates:
        # Block if candidate has marked/approved fees
        has_locked_fees = CandidateFee.objects.filter(
            candidate=candidate,
            verification_status__in=['marked', 'approved']
        ).exists()
        if has_locked_fees:
            skipped.append({
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has fees marked/approved by accounts'
            })
            continue
        
        # Delete all modular results
        modular_deleted = ModularResult.objects.filter(candidate=candidate).delete()
        total_cleared['modular_results'] += modular_deleted[0] if modular_deleted else 0
        
        # Delete all formal results
        formal_deleted = FormalResult.objects.filter(candidate=candidate).delete()
        total_cleared['formal_results'] += formal_deleted[0] if formal_deleted else 0
        
        # Delete all workers PAS results
        workers_deleted = WorkersPasResult.objects.filter(candidate=candidate).delete()
        total_cleared['workers_pas_results'] += workers_deleted[0] if workers_deleted else 0
        
        # Delete all enrollments
        enrollments_deleted = CandidateEnrollment.objects.filter(candidate=candidate).delete()
        total_cleared['enrollments'] += enrollments_deleted[0] if enrollments_deleted else 0
        
        total_cleared['candidates_processed'] += 1
    
    return Response({
        'message': f'Successfully cleared data for {total_cleared["candidates_processed"]} candidate(s)',
        'cleared': total_cleared,
        'skipped': skipped
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_change_candidate_center(request):
    """Bulk change assessment center for multiple candidates"""
    if request.user.is_authenticated and request.user.user_type == 'center_representative':
        return Response({'error': 'Permission denied. Center representatives cannot perform this action.'}, status=status.HTTP_403_FORBIDDEN)

    candidate_ids = request.data.get('candidate_ids', [])
    new_center_id = request.data.get('new_center_id')
    
    if not new_center_id:
        return Response(
            {'error': 'New assessment center ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not candidate_ids:
        return Response(
            {'error': 'No candidates selected'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from assessment_centers.models import AssessmentCenter
    try:
        new_center = AssessmentCenter.objects.get(id=new_center_id)
    except AssessmentCenter.DoesNotExist:
        return Response(
            {'error': 'Assessment center not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from fees.models import CandidateFee, CenterFee
    
    candidates = Candidate.objects.filter(id__in=candidate_ids)
    
    total_updated = {
        'candidates': 0,
        'fees_moved': 0,
    }
    failed = []
    
    for candidate in candidates:
        try:
            old_center = candidate.assessment_center
            
            # Update candidate's assessment center
            candidate.assessment_center = new_center
            candidate.assessment_center_branch = None  # Reset branch
            
            # Generate new registration number if candidate is submitted
            if candidate.is_submitted and candidate.registration_number:
                new_registration_number = candidate.generate_registration_number()
                candidate.registration_number = new_registration_number
                
                # Regenerate payment code
                new_payment_code = candidate.generate_payment_code()
                candidate.payment_code = new_payment_code
            
            candidate.save()
            
            # Update CenterFee totals if candidate has fees
            candidate_fees = CandidateFee.objects.filter(candidate=candidate)
            
            for candidate_fee in candidate_fees:
                if candidate_fee.total_amount > 0:
                    total_updated['fees_moved'] += 1
                    
                    # Decrease old center's fee totals
                    if old_center:
                        try:
                            old_center_fee = CenterFee.objects.get(
                                assessment_center=old_center,
                                assessment_series=candidate_fee.assessment_series
                            )
                            old_center_fee.total_candidates = max(0, old_center_fee.total_candidates - 1)
                            old_center_fee.total_amount = max(0, old_center_fee.total_amount - candidate_fee.total_amount)
                            old_center_fee.save()
                        except CenterFee.DoesNotExist:
                            pass
                    
                    # Increase new center's fee totals
                    new_center_fee, created = CenterFee.objects.get_or_create(
                        assessment_center=new_center,
                        assessment_series=candidate_fee.assessment_series,
                        defaults={
                            'total_candidates': 0,
                            'total_amount': 0,
                            'amount_due': 0,
                        }
                    )
                    new_center_fee.total_candidates += 1
                    new_center_fee.total_amount += candidate_fee.total_amount
                    new_center_fee.save()
            
            total_updated['candidates'] += 1
        except Exception as e:
            failed.append({
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reason': str(e)
            })
    
    return Response({
        'message': f'Successfully moved {total_updated["candidates"]} candidate(s) to {new_center.center_name}',
        'new_center': {
            'id': new_center.id,
            'name': new_center.center_name,
        },
        'updated': total_updated,
        'failed': failed,
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def clear_candidate_data(request, candidate_id):
    """Clear all results, enrollments, and fees for a candidate"""
    if request.user.is_authenticated and request.user.user_type == 'center_representative':
        return Response({'error': 'Permission denied. Center representatives cannot perform this action.'}, status=status.HTTP_403_FORBIDDEN)

    try:
        candidate = Candidate.objects.get(id=candidate_id)
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    from fees.models import CandidateFee
    
    # Block if candidate has marked/approved fees
    has_locked_fees = CandidateFee.objects.filter(
        candidate=candidate,
        verification_status__in=['marked', 'approved']
    ).exists()
    if has_locked_fees:
        return Response(
            {'error': 'Cannot clear data: candidate has fees marked/approved by accounts. Contact accounts office first.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    cleared = {
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
        'enrollments': 0,
    }
    
    # Delete all modular results
    modular_deleted = ModularResult.objects.filter(candidate=candidate).delete()
    cleared['modular_results'] = modular_deleted[0] if modular_deleted else 0
    
    # Delete all formal results
    formal_deleted = FormalResult.objects.filter(candidate=candidate).delete()
    cleared['formal_results'] = formal_deleted[0] if formal_deleted else 0
    
    # Delete all workers PAS results
    workers_deleted = WorkersPasResult.objects.filter(candidate=candidate).delete()
    cleared['workers_pas_results'] = workers_deleted[0] if workers_deleted else 0
    
    # Delete all enrollments (this will also reset fees since fees are tied to enrollments)
    enrollments_deleted = CandidateEnrollment.objects.filter(candidate=candidate).delete()
    cleared['enrollments'] = enrollments_deleted[0] if enrollments_deleted else 0

    actor = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
    CandidateActivity.objects.create(
        candidate=candidate,
        actor=actor,
        action='candidate_data_cleared',
        description='Candidate results, enrollments & fees cleared',
        details={'cleared': cleared},
    )
    
    return Response({
        'message': f'Successfully cleared all data for {candidate.full_name}',
        'candidate_id': candidate.id,
        'candidate_name': candidate.full_name,
        'cleared': cleared
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def candidate_login(request):
    """
    Authenticate a candidate using their registration number.
    Returns candidate data if registration number exists.
    """
    registration_number = request.data.get('registration_number', '').strip().upper()
    
    if not registration_number:
        return Response(
            {'error': 'Registration number is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        candidate = Candidate.objects.select_related(
            'occupation', 'assessment_center', 'assessment_center_branch'
        ).get(registration_number=registration_number)
        
        # Generate a simple token for the candidate session
        import hashlib
        import time
        token = hashlib.sha256(f"{candidate.id}{time.time()}".encode()).hexdigest()[:32]
        
        # Return candidate basic info
        candidate_data = {
            'id': candidate.id,
            'registration_number': candidate.registration_number,
            'full_name': candidate.full_name,
            'photo': candidate.passport_photo.url if candidate.passport_photo else None,
            'gender': candidate.gender,
            'date_of_birth': candidate.date_of_birth,
            'nationality': candidate.nationality,
            'registration_category': candidate.registration_category,
            'registration_category_display': candidate.get_registration_category_display(),
            'occupation': {
                'id': candidate.occupation.id,
                'name': candidate.occupation.occ_name,
                'code': candidate.occupation.occ_code,
            } if candidate.occupation else None,
            'assessment_center': {
                'id': candidate.assessment_center.id,
                'name': candidate.assessment_center.center_name,
            } if candidate.assessment_center else None,
            'status': candidate.status,
            'is_verified': candidate.status == 'verified',
        }
        
        return Response({
            'token': token,
            'candidate': candidate_data
        }, status=status.HTTP_200_OK)
        
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Registration number not found. Please check and try again.'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def candidate_portal_data(request, registration_number):
    """
    Get full candidate portal data including bio, enrollments, and results.
    Results show only 'Uploaded' status, not raw marks.
    """
    try:
        candidate = Candidate.objects.select_related(
            'occupation', 'assessment_center', 'assessment_center_branch'
        ).get(registration_number=registration_number.upper())
        
        # Bio Data
        bio_data = {
            'registration_number': candidate.registration_number,
            'payment_code': candidate.payment_code,
            'full_name': candidate.full_name,
            'photo': candidate.passport_photo.url if candidate.passport_photo else None,
            'gender': candidate.gender,
            'date_of_birth': candidate.date_of_birth,
            'nationality': candidate.nationality,
            'contact': candidate.contact,
            'district': candidate.district.name if candidate.district else None,
            'village': candidate.village.name if candidate.village else None,
            'is_refugee': candidate.is_refugee,
            'refugee_number': candidate.refugee_number,
            'has_disability': candidate.has_disability,
            'disability': candidate.nature_of_disability.name if candidate.nature_of_disability else None,
        }
        
        # Occupation Info
        occupation_info = {
            'registration_category': candidate.get_registration_category_display(),
            'occupation': candidate.occupation.occ_name if candidate.occupation else None,
            'occupation_code': candidate.occupation.occ_code if candidate.occupation else None,
            'assessment_center': candidate.assessment_center.center_name if candidate.assessment_center else None,
            'entry_year': candidate.entry_year,
            'intake': candidate.get_intake_display() if candidate.intake else None,
            'preferred_language': candidate.preferred_assessment_language,
        }
        
        # Enrollments
        enrollments = CandidateEnrollment.objects.filter(
            candidate=candidate,
            is_active=True
        ).select_related('assessment_series', 'occupation_level')
        
        enrollment_data = []
        for enrollment in enrollments:
            enrollment_data.append({
                'id': enrollment.id,
                'assessment_series': enrollment.assessment_series.name,
                'level': enrollment.occupation_level.level_name if enrollment.occupation_level else None,
                'enrolled_date': enrollment.enrolled_at,
                'total_amount': float(enrollment.total_amount) if enrollment.total_amount else 0,
            })
        
        # Results (only from released assessment series, show 'Uploaded' instead of marks)
        results_data = []
        
        if candidate.registration_category == 'modular':
            from results.models import ModularResult
            results = ModularResult.objects.filter(
                candidate=candidate,
                assessment_series__results_released=True
            ).select_related('assessment_series', 'module')
            
            for result in results:
                results_data.append({
                    'assessment_series': result.assessment_series.name,
                    'module': result.module.module_name,
                    'module_code': result.module.module_code,
                    'type': result.get_type_display(),
                    'mark_status': 'Uploaded' if result.mark is not None else 'Pending',
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                })
        
        elif candidate.registration_category == 'formal':
            from results.models import FormalResult
            results = FormalResult.objects.filter(
                candidate=candidate,
                assessment_series__results_released=True
            ).select_related('assessment_series', 'level', 'exam', 'paper')
            
            for result in results:
                exam_or_paper = ''
                if result.exam:
                    exam_or_paper = result.exam.module_name
                elif result.paper:
                    exam_or_paper = result.paper.paper_name
                
                results_data.append({
                    'assessment_series': result.assessment_series.name,
                    'level': result.level.level_name,
                    'exam_or_paper': exam_or_paper,
                    'type': result.get_type_display(),
                    'mark_status': 'Uploaded' if result.mark is not None else 'Pending',
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                })
        
        elif candidate.registration_category == 'workers_pas':
            from results.models import WorkersPasResult
            results = WorkersPasResult.objects.filter(
                candidate=candidate,
                assessment_series__results_released=True
            ).select_related('assessment_series', 'level', 'module', 'paper')
            
            for result in results:
                results_data.append({
                    'assessment_series': result.assessment_series.name,
                    'level': result.level.level_name,
                    'module': result.module.module_name,
                    'paper': result.paper.paper_name,
                    'type': 'Practical',
                    'mark_status': 'Uploaded' if result.mark is not None else 'Pending',
                    'grade': result.grade,
                    'comment': result.comment,
                    'status': result.get_status_display(),
                })
        
        return Response({
            'bio_data': bio_data,
            'occupation_info': occupation_info,
            'enrollments': enrollment_data,
            'results': results_data,
            'status': candidate.status,
            'is_verified': candidate.status == 'verified',
        }, status=status.HTTP_200_OK)
        
    except Candidate.DoesNotExist:
        return Response(
            {'error': 'Candidate not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def enrollment_list_view(request):
    """
    List all enrollments with filtering options
    """
    from emis.pagination import FlexiblePagination
    
    queryset = CandidateEnrollment.objects.select_related(
        'candidate',
        'candidate__assessment_center',
        'candidate__occupation',
        'assessment_series',
        'occupation_level',
        'occupation_level__occupation'
    ).prefetch_related(
        'modules',
        'modules__module',
        'modules__module__occupation',
        'papers',
        'papers__paper',
        'papers__paper__occupation',
        'papers__paper__level'
    ).filter(is_active=True).order_by('-enrolled_at')
    
    # Filter by registration category
    registration_category = request.query_params.get('registration_category')
    if registration_category:
        queryset = queryset.filter(candidate__registration_category=registration_category)
    
    # Filter by assessment series
    assessment_series = request.query_params.get('assessment_series')
    if assessment_series:
        queryset = queryset.filter(assessment_series_id=assessment_series)
    
    # Filter by assessment center
    assessment_center = request.query_params.get('assessment_center')
    if assessment_center:
        queryset = queryset.filter(candidate__assessment_center_id=assessment_center)
    
    # Filter by occupation
    occupation = request.query_params.get('occupation')
    if occupation:
        queryset = queryset.filter(
            Q(occupation_level__occupation_id=occupation) |
            Q(candidate__occupation_id=occupation)
        )
    
    # Search by registration number or name
    search = request.query_params.get('search')
    if search:
        queryset = queryset.filter(
            Q(candidate__registration_number__icontains=search) |
            Q(candidate__full_name__icontains=search)
        )
    
    # Filter by center for center representatives
    if request.user.is_authenticated and request.user.user_type == 'center_representative':
        if hasattr(request.user, 'center_rep_profile'):
            center_rep = request.user.center_rep_profile
            queryset = queryset.filter(candidate__assessment_center=center_rep.assessment_center)
            if center_rep.assessment_center_branch:
                queryset = queryset.filter(candidate__assessment_center_branch=center_rep.assessment_center_branch)
    
    # Paginate
    paginator = FlexiblePagination()
    page = paginator.paginate_queryset(queryset, request)
    
    if page is not None:
        serializer = EnrollmentListSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)
    
    serializer = EnrollmentListSerializer(queryset, many=True)
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_change_enrollment_series(request):
    """Bulk change assessment series for specific enrollments"""
    enrollment_ids = request.data.get('enrollment_ids', [])
    new_series_id = request.data.get('new_series_id')
    select_all = request.data.get('select_all', False)
    filters = request.data.get('filters', {})
    
    if not new_series_id:
        return Response(
            {'error': 'New assessment series ID is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    from assessment_series.models import AssessmentSeries
    try:
        new_series = AssessmentSeries.objects.get(id=new_series_id)
    except AssessmentSeries.DoesNotExist:
        return Response(
            {'error': 'Assessment series not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    
    # Get enrollments to update
    if select_all:
        # Apply filters to get all matching enrollments
        queryset = CandidateEnrollment.objects.select_related(
            'candidate', 'assessment_series', 'occupation_level'
        ).all()
        
        if filters.get('registration_category'):
            queryset = queryset.filter(candidate__registration_category=filters['registration_category'])
        if filters.get('assessment_series'):
            queryset = queryset.filter(assessment_series_id=filters['assessment_series'])
        if filters.get('assessment_center'):
            queryset = queryset.filter(candidate__assessment_center_id=filters['assessment_center'])
        if filters.get('occupation'):
            queryset = queryset.filter(
                Q(occupation_level__occupation_id=filters['occupation']) |
                Q(candidate__occupation_id=filters['occupation'])
            )
        if filters.get('search'):
            queryset = queryset.filter(
                Q(candidate__registration_number__icontains=filters['search']) |
                Q(candidate__full_name__icontains=filters['search'])
            )
        enrollments = queryset
    else:
        if not enrollment_ids:
            return Response(
                {'error': 'No enrollments selected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        enrollments = CandidateEnrollment.objects.filter(id__in=enrollment_ids)
    
    total_updated = {
        'enrollments': 0,
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
    }
    
    for enrollment in enrollments:
        candidate = enrollment.candidate
        old_series = enrollment.assessment_series
        
        # Update the enrollment
        enrollment.assessment_series = new_series
        enrollment.save()
        total_updated['enrollments'] += 1
        
        # Update results for this candidate from old series to new series
        modular_updated = ModularResult.objects.filter(
            candidate=candidate, assessment_series=old_series
        ).update(assessment_series=new_series)
        total_updated['modular_results'] += modular_updated
        
        formal_updated = FormalResult.objects.filter(
            candidate=candidate, assessment_series=old_series
        ).update(assessment_series=new_series)
        total_updated['formal_results'] += formal_updated
        
        workers_updated = WorkersPasResult.objects.filter(
            candidate=candidate, assessment_series=old_series
        ).update(assessment_series=new_series)
        total_updated['workers_pas_results'] += workers_updated
    
    return Response({
        'message': f'Successfully moved {total_updated["enrollments"]} enrollment(s) to {new_series.name}',
        'new_series': {
            'id': new_series.id,
            'name': new_series.name,
        },
        'updated': total_updated
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_de_enroll_by_enrollment(request):
    """Bulk de-enroll by enrollment IDs - deletes enrollments and clears fees"""
    enrollment_ids = request.data.get('enrollment_ids', [])
    select_all = request.data.get('select_all', False)
    filters = request.data.get('filters', {})
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    from fees.models import CandidateFee
    
    # Get enrollments to delete
    if select_all:
        queryset = CandidateEnrollment.objects.select_related(
            'candidate', 'assessment_series', 'occupation_level'
        ).all()
        
        if filters.get('registration_category'):
            queryset = queryset.filter(candidate__registration_category=filters['registration_category'])
        if filters.get('assessment_series'):
            queryset = queryset.filter(assessment_series_id=filters['assessment_series'])
        if filters.get('assessment_center'):
            queryset = queryset.filter(candidate__assessment_center_id=filters['assessment_center'])
        if filters.get('occupation'):
            queryset = queryset.filter(
                Q(occupation_level__occupation_id=filters['occupation']) |
                Q(candidate__occupation_id=filters['occupation'])
            )
        if filters.get('search'):
            queryset = queryset.filter(
                Q(candidate__registration_number__icontains=filters['search']) |
                Q(candidate__full_name__icontains=filters['search'])
            )
        enrollments = list(queryset)
    else:
        if not enrollment_ids:
            return Response(
                {'error': 'No enrollments selected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        enrollments = list(CandidateEnrollment.objects.filter(id__in=enrollment_ids))
    
    success_count = 0
    fees_deleted = 0
    failed = []
    skipped_with_marks = []
    
    for enrollment in enrollments:
        candidate = enrollment.candidate
        series = enrollment.assessment_series
        
        # Check if candidate has any results for this enrollment
        has_modular_results = ModularResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).exists()
        
        has_formal_results = FormalResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).exists()
        
        has_workers_pas_results = WorkersPasResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).exists()
        
        if has_modular_results or has_formal_results or has_workers_pas_results:
            skipped_with_marks.append({
                'enrollment_id': enrollment.id,
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has marks - cannot de-enroll'
            })
            continue
        
        # Check if candidate has marked/approved fees for this series
        has_locked_fees = CandidateFee.objects.filter(
            candidate=candidate, assessment_series=series,
            verification_status__in=['marked', 'approved']
        ).exists()
        if has_locked_fees:
            skipped_with_marks.append({
                'enrollment_id': enrollment.id,
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has fees marked/approved by accounts - cannot de-enroll'
            })
            continue
        
        try:
            # Delete fees for this candidate and series
            deleted_fees = CandidateFee.objects.filter(
                candidate=candidate, assessment_series=series
            ).delete()[0]
            fees_deleted += deleted_fees
            
            # Delete the enrollment
            enrollment.delete()
            success_count += 1
        except Exception as e:
            failed.append({
                'enrollment_id': enrollment.id,
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reason': str(e)
            })
    
    return Response({
        'message': f'Successfully de-enrolled {success_count} candidate(s)',
        'success_count': success_count,
        'fees_deleted': fees_deleted,
        'skipped_with_marks': skipped_with_marks,
        'failed': failed
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_clear_enrollment_data(request):
    """Bulk clear results, enrollments, and fees for specific enrollments"""
    if request.user.is_authenticated and request.user.user_type == 'center_representative':
        return Response({'error': 'Permission denied. Center representatives cannot perform this action.'}, status=status.HTTP_403_FORBIDDEN)

    enrollment_ids = request.data.get('enrollment_ids', [])
    select_all = request.data.get('select_all', False)
    filters = request.data.get('filters', {})
    
    from results.models import ModularResult, FormalResult, WorkersPasResult
    from fees.models import CandidateFee
    
    # Get enrollments to clear
    if select_all:
        queryset = CandidateEnrollment.objects.select_related(
            'candidate', 'assessment_series', 'occupation_level'
        ).all()
        
        if filters.get('registration_category'):
            queryset = queryset.filter(candidate__registration_category=filters['registration_category'])
        if filters.get('assessment_series'):
            queryset = queryset.filter(assessment_series_id=filters['assessment_series'])
        if filters.get('assessment_center'):
            queryset = queryset.filter(candidate__assessment_center_id=filters['assessment_center'])
        if filters.get('occupation'):
            queryset = queryset.filter(
                Q(occupation_level__occupation_id=filters['occupation']) |
                Q(candidate__occupation_id=filters['occupation'])
            )
        if filters.get('search'):
            queryset = queryset.filter(
                Q(candidate__registration_number__icontains=filters['search']) |
                Q(candidate__full_name__icontains=filters['search'])
            )
        enrollments = list(queryset)
    else:
        if not enrollment_ids:
            return Response(
                {'error': 'No enrollments selected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        enrollments = list(CandidateEnrollment.objects.filter(id__in=enrollment_ids))
    
    total_cleared = {
        'modular_results': 0,
        'formal_results': 0,
        'workers_pas_results': 0,
        'enrollments': 0,
        'fees': 0,
    }
    skipped = []
    
    for enrollment in enrollments:
        candidate = enrollment.candidate
        series = enrollment.assessment_series
        
        # Block if candidate has marked/approved fees for this series
        has_locked_fees = CandidateFee.objects.filter(
            candidate=candidate, assessment_series=series,
            verification_status__in=['marked', 'approved']
        ).exists()
        if has_locked_fees:
            skipped.append({
                'enrollment_id': enrollment.id,
                'candidate_id': candidate.id,
                'name': candidate.full_name,
                'reg_no': candidate.registration_number,
                'reason': 'Has fees marked/approved by accounts'
            })
            continue
        
        # Delete results for this candidate and series
        modular_deleted = ModularResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).delete()[0]
        total_cleared['modular_results'] += modular_deleted
        
        formal_deleted = FormalResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).delete()[0]
        total_cleared['formal_results'] += formal_deleted
        
        workers_deleted = WorkersPasResult.objects.filter(
            candidate=candidate, assessment_series=series
        ).delete()[0]
        total_cleared['workers_pas_results'] += workers_deleted
        
        # Delete fees for this candidate and series
        fees_deleted = CandidateFee.objects.filter(
            candidate=candidate, assessment_series=series
        ).delete()[0]
        total_cleared['fees'] += fees_deleted
        
        # Delete the enrollment
        enrollment.delete()
        total_cleared['enrollments'] += 1
    
    return Response({
        'message': f'Successfully cleared data for {total_cleared["enrollments"]} enrollment(s)',
        'cleared': total_cleared,
        'skipped': skipped
    }, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_update_enrollment(request):
    """Bulk update enrollments - set level for formal, modules for modular, papers for workers_pas"""
    enrollment_ids = request.data.get('enrollment_ids', [])
    select_all = request.data.get('select_all', False)
    filters = request.data.get('filters', {})
    
    # Update data
    level_id = request.data.get('level_id')  # For formal candidates
    module_ids = request.data.get('module_ids', [])  # For modular candidates
    paper_ids = request.data.get('paper_ids', [])  # For workers_pas candidates
    
    from occupations.models import OccupationLevel, OccupationModule, OccupationPaper
    
    # Get enrollments to update
    if select_all:
        queryset = CandidateEnrollment.objects.select_related(
            'candidate', 'assessment_series', 'occupation_level'
        ).all()
        
        if filters.get('registration_category'):
            queryset = queryset.filter(candidate__registration_category=filters['registration_category'])
        if filters.get('assessment_series'):
            queryset = queryset.filter(assessment_series_id=filters['assessment_series'])
        if filters.get('assessment_center'):
            queryset = queryset.filter(candidate__assessment_center_id=filters['assessment_center'])
        if filters.get('occupation'):
            queryset = queryset.filter(
                Q(occupation_level__occupation_id=filters['occupation']) |
                Q(candidate__occupation_id=filters['occupation'])
            )
        if filters.get('search'):
            queryset = queryset.filter(
                Q(candidate__registration_number__icontains=filters['search']) |
                Q(candidate__full_name__icontains=filters['search'])
            )
        enrollments = list(queryset)
    else:
        if not enrollment_ids:
            return Response(
                {'error': 'No enrollments selected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        enrollments = list(CandidateEnrollment.objects.select_related('candidate').filter(id__in=enrollment_ids))
    
    # Validate level if provided
    level = None
    if level_id:
        try:
            level = OccupationLevel.objects.get(id=level_id)
        except OccupationLevel.DoesNotExist:
            return Response({'error': 'Level not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Validate modules if provided
    modules = []
    if module_ids:
        modules = list(OccupationModule.objects.filter(id__in=module_ids))
        if len(modules) != len(module_ids):
            return Response({'error': 'Some modules not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Validate papers if provided
    papers = []
    if paper_ids:
        papers = list(OccupationPaper.objects.filter(id__in=paper_ids))
        if len(papers) != len(paper_ids):
            return Response({'error': 'Some papers not found'}, status=status.HTTP_404_NOT_FOUND)
    
    updated_count = 0
    skipped = []
    
    for enrollment in enrollments:
        candidate = enrollment.candidate
        reg_category = candidate.registration_category
        
        try:
            if reg_category == 'formal' and level:
                # Update level for formal candidates
                enrollment.occupation_level = level
                enrollment.save()
                updated_count += 1
                
            elif reg_category == 'modular' and modules:
                # Clear existing modules and add new ones
                EnrollmentModule.objects.filter(enrollment=enrollment).delete()
                for module in modules:
                    EnrollmentModule.objects.create(enrollment=enrollment, module=module)
                # Set the level from the first module
                if modules:
                    enrollment.occupation_level = modules[0].level
                    enrollment.save()
                updated_count += 1
                
            elif reg_category == 'workers_pas' and papers:
                # Clear existing papers and add new ones
                EnrollmentPaper.objects.filter(enrollment=enrollment).delete()
                for paper in papers:
                    EnrollmentPaper.objects.create(enrollment=enrollment, paper=paper)
                updated_count += 1
                
            else:
                skipped.append({
                    'enrollment_id': enrollment.id,
                    'candidate': candidate.full_name,
                    'reason': f'No matching update data for {reg_category} category'
                })
        except Exception as e:
            skipped.append({
                'enrollment_id': enrollment.id,
                'candidate': candidate.full_name,
                'reason': str(e)
            })
    
    return Response({
        'message': f'Successfully updated {updated_count} enrollment(s)',
        'updated_count': updated_count,
        'skipped': skipped
    }, status=status.HTTP_200_OK)
