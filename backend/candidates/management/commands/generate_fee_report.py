"""
Management command to generate fee report Excel workbook.

Calculates fees based on enrollment data for candidates who may not have fees set.

Fee Structure:
- Formal: Level 1 = 80,000, Level 2 = 100,000, Level 3 = 150,000, Level 4 = 220,000
- Modular: 1 module = 70,000, 2 modules = 90,000
- Worker's PAS: 2 papers = 150,000, 3 papers = 225,000, 4 papers = 300,000 (min 2, max 4 papers)

Usage:
    python manage.py generate_fee_report
    python manage.py generate_fee_report --series "August 2025"
    python manage.py generate_fee_report --output /path/to/report.xlsx
"""

import re
from datetime import datetime
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from candidates.models import Candidate, CandidateEnrollment
from assessment_series.models import AssessmentSeries


class Command(BaseCommand):
    help = 'Generate fee report Excel workbook based on enrollment data'

    # Fee structure
    FORMAL_FEES = {
        1: 80000,   # Level 1
        2: 100000,  # Level 2
        3: 150000,  # Level 3
        4: 220000,  # Level 4
    }
    
    MODULAR_FEES = {
        1: 70000,   # 1 module
        2: 90000,   # 2 modules
    }
    
    WORKERS_PAS_FEES = {
        2: 150000,  # 2 papers
        3: 225000,  # 3 papers
        4: 300000,  # 4 papers
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--series',
            type=str,
            help='Filter by assessment series name (optional)',
        )
        parser.add_argument(
            '--output',
            type=str,
            help='Output file path (default: fee_report_YYYYMMDD_HHMMSS.xlsx)',
        )

    def handle(self, *args, **options):
        series_filter = options.get('series')
        output_path = options.get('output')
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'/tmp/fee_report_{timestamp}.xlsx'
        
        self.stdout.write(self.style.NOTICE('Generating fee report...'))
        
        # Get enrollments
        enrollments = CandidateEnrollment.objects.select_related(
            'candidate',
            'candidate__assessment_center',
            'assessment_series',
            'occupation_level'
        ).prefetch_related(
            'modules',
            'papers'
        )
        
        if series_filter:
            enrollments = enrollments.filter(assessment_series__name__icontains=series_filter)
            self.stdout.write(f'Filtering by series: {series_filter}')
        
        # Process enrollments - group Worker's PAS by candidate+series
        formal_data = []
        modular_data = []
        workers_pas_data = []
        
        # Track Worker's PAS candidates to combine papers across enrollments
        # Key: (candidate_id, series_id) -> {candidate info, total_papers set}
        workers_pas_grouped = defaultdict(lambda: {
            'candidate': None,
            'center_name': None,
            'series_name': None,
            'papers': set(),  # Use set to avoid duplicates
        })
        
        for enrollment in enrollments:
            candidate = enrollment.candidate
            center_name = candidate.assessment_center.center_name if candidate.assessment_center else 'N/A'
            series_name = enrollment.assessment_series.name
            series_id = enrollment.assessment_series.id
            reg_category = candidate.registration_category
            
            if reg_category == 'formal':
                fee = self.calculate_formal_fee(enrollment)
                row = {
                    'assessment_center': center_name,
                    'assessment_series': series_name,
                    'registration_number': candidate.registration_number,
                    'candidate_name': candidate.full_name,
                    'fee': fee,
                    'details': enrollment.occupation_level.level_name if enrollment.occupation_level else "No level",
                }
                formal_data.append(row)
                
            elif reg_category == 'modular':
                module_count = enrollment.modules.count()
                fee = self.MODULAR_FEES.get(module_count, 0)
                row = {
                    'assessment_center': center_name,
                    'assessment_series': series_name,
                    'registration_number': candidate.registration_number,
                    'candidate_name': candidate.full_name,
                    'fee': fee,
                    'details': f"{module_count} module(s)",
                }
                modular_data.append(row)
                
            elif reg_category == 'workers_pas':
                # Group by candidate + series, collect all papers
                key = (candidate.id, series_id)
                workers_pas_grouped[key]['candidate'] = candidate
                workers_pas_grouped[key]['center_name'] = center_name
                workers_pas_grouped[key]['series_name'] = series_name
                
                # Add all papers from this enrollment
                for paper in enrollment.papers.all():
                    workers_pas_grouped[key]['papers'].add(paper.id)
        
        # Process grouped Worker's PAS data
        for key, data in workers_pas_grouped.items():
            candidate = data['candidate']
            paper_count = len(data['papers'])
            fee = self.WORKERS_PAS_FEES.get(paper_count, 0)
            
            row = {
                'assessment_center': data['center_name'],
                'assessment_series': data['series_name'],
                'registration_number': candidate.registration_number,
                'candidate_name': candidate.full_name,
                'fee': fee,
                'details': f"{paper_count} paper(s)",
            }
            workers_pas_data.append(row)
        
        # Combine all data for the "All" sheet
        all_data = formal_data + modular_data + workers_pas_data
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: All
        ws_all = wb.active
        ws_all.title = 'All'
        self.write_sheet(ws_all, all_data, 'All Candidates Fee Report')
        
        # Sheet 2: Formal
        ws_formal = wb.create_sheet('Formal')
        self.write_sheet(ws_formal, formal_data, 'Formal Candidates Fee Report')
        
        # Sheet 3: Modular
        ws_modular = wb.create_sheet('Modular')
        self.write_sheet(ws_modular, modular_data, 'Modular Candidates Fee Report')
        
        # Sheet 4: Worker's PAS
        ws_workers = wb.create_sheet("Worker's PAS")
        self.write_sheet(ws_workers, workers_pas_data, "Worker's PAS Candidates Fee Report")
        
        # Sheet 5: Center Fee Summary
        ws_center_summary = wb.create_sheet("Center Fee Summary")
        center_summary_data = self.get_center_summary_data(all_data)
        self.write_center_summary_sheet(ws_center_summary, center_summary_data)
        
        # Save workbook
        wb.save(output_path)
        
        # Print summary
        self.stdout.write(self.style.SUCCESS(f'\nReport saved to: {output_path}'))
        self.stdout.write(f'\nSummary:')
        self.stdout.write(f'  Total candidates: {len(all_data)}')
        self.stdout.write(f'  Formal: {len(formal_data)} (Total: {sum(r["fee"] for r in formal_data):,} UGX)')
        self.stdout.write(f'  Modular: {len(modular_data)} (Total: {sum(r["fee"] for r in modular_data):,} UGX)')
        self.stdout.write(f"  Worker's PAS: {len(workers_pas_data)} (Total: {sum(r['fee'] for r in workers_pas_data):,} UGX)")
        self.stdout.write(f'\n  GRAND TOTAL: {sum(r["fee"] for r in all_data):,} UGX')
        
        # Print per-center summary
        self.print_center_summary(all_data)

    def calculate_formal_fee(self, enrollment):
        """Calculate fee for formal candidates based on level"""
        if enrollment.occupation_level:
            level_name = enrollment.occupation_level.level_name
            match = re.search(r'\d+', level_name)
            level_num = int(match.group()) if match else 0
            return self.FORMAL_FEES.get(level_num, 0)
        return 0

    def get_center_summary_data(self, all_data):
        """Generate center fee summary grouped by center and series"""
        # Group by (center, series)
        summary = defaultdict(lambda: {'candidates': 0, 'fees': 0})
        
        for row in all_data:
            key = (row['assessment_center'], row['assessment_series'])
            summary[key]['candidates'] += 1
            summary[key]['fees'] += row['fee']
        
        # Convert to list for writing
        result = []
        for (center, series), data in sorted(summary.items()):
            result.append({
                'center': center,
                'series': series,
                'candidates': data['candidates'],
                'fees': data['fees'],
            })
        
        return result

    def write_sheet(self, ws, data, title):
        """Write data to a worksheet with formatting"""
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = '#,##0'
        
        # Title row
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Assessment Center', 'Assessment Series', 'Reg No', 'Candidate Name', 'Details', 'Fees (UGX)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, row_data in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=row_data['assessment_center']).border = border
            ws.cell(row=row_idx, column=2, value=row_data['assessment_series']).border = border
            ws.cell(row=row_idx, column=3, value=row_data['registration_number']).border = border
            ws.cell(row=row_idx, column=4, value=row_data['candidate_name']).border = border
            ws.cell(row=row_idx, column=5, value=row_data['details']).border = border
            fee_cell = ws.cell(row=row_idx, column=6, value=row_data['fee'])
            fee_cell.border = border
            fee_cell.number_format = money_format
            fee_cell.alignment = Alignment(horizontal='right')
        
        # Total row
        if data:
            total_row = len(data) + 4
            ws.cell(row=total_row, column=5, value='TOTAL:').font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=6, value=sum(r['fee'] for r in data))
            total_cell.font = Font(bold=True)
            total_cell.number_format = money_format
            total_cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        column_widths = [30, 25, 30, 35, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def write_center_summary_sheet(self, ws, data):
        """Write center fee summary sheet"""
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = '#,##0'
        
        # Title row
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Center Fee Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Assessment Center', 'Assessment Series', 'Candidates', 'Fees (UGX)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, row_data in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=row_data['center']).border = border
            ws.cell(row=row_idx, column=2, value=row_data['series']).border = border
            
            candidates_cell = ws.cell(row=row_idx, column=3, value=row_data['candidates'])
            candidates_cell.border = border
            candidates_cell.alignment = Alignment(horizontal='center')
            
            fee_cell = ws.cell(row=row_idx, column=4, value=row_data['fees'])
            fee_cell.border = border
            fee_cell.number_format = money_format
            fee_cell.alignment = Alignment(horizontal='right')
        
        # Total row
        if data:
            total_row = len(data) + 4
            ws.cell(row=total_row, column=2, value='TOTAL:').font = Font(bold=True)
            
            total_candidates = ws.cell(row=total_row, column=3, value=sum(r['candidates'] for r in data))
            total_candidates.font = Font(bold=True)
            total_candidates.alignment = Alignment(horizontal='center')
            
            total_fees = ws.cell(row=total_row, column=4, value=sum(r['fees'] for r in data))
            total_fees.font = Font(bold=True)
            total_fees.number_format = money_format
            total_fees.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        column_widths = [35, 25, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def print_center_summary(self, data):
        """Print summary grouped by assessment center"""
        if not data:
            return
        
        self.stdout.write('\n--- Per Assessment Center Summary ---')
        
        # Group by center
        center_totals = {}
        for row in data:
            center = row['assessment_center']
            if center not in center_totals:
                center_totals[center] = {'count': 0, 'total': 0}
            center_totals[center]['count'] += 1
            center_totals[center]['total'] += row['fee']
        
        # Sort by total descending
        sorted_centers = sorted(center_totals.items(), key=lambda x: x[1]['total'], reverse=True)
        
        for center, info in sorted_centers:
            self.stdout.write(f"  {center}: {info['count']} candidates, {info['total']:,} UGX")
