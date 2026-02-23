from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Count
from django.http import HttpResponse
from django.conf import settings
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.pdfgen import canvas
from datetime import datetime
import os
from candidates.models import Candidate
from assessment_series.models import AssessmentSeries
from assessment_centers.models import AssessmentCenter
from .signals import update_center_fee
from .models import CandidateFee, CenterFee
from .serializers import CandidateFeeSerializer, CenterFeeSerializer


class CandidateFeeViewSet(viewsets.ModelViewSet):
    """ViewSet for managing candidate fees"""
    serializer_class = CandidateFeeSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['assessment_series', 'payment_status', 'attempt_status', 'verification_status', 'candidate', 'candidate__occupation']
    search_fields = ['payment_code', 'candidate__registration_number', 'candidate__first_name', 'candidate__last_name']
    ordering_fields = ['created_at', 'total_amount', 'payment_date']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Filter fees by center for center representatives"""
        queryset = CandidateFee.objects.select_related(
            'candidate', 'candidate__occupation', 'assessment_series',
            'marked_by', 'approved_by'
        ).all()
        
        # Filter by center for center representatives
        if self.request.user.is_authenticated and self.request.user.user_type == 'center_representative':
            if hasattr(self.request.user, 'center_rep_profile'):
                center_rep = self.request.user.center_rep_profile
                queryset = queryset.filter(candidate__assessment_center=center_rep.assessment_center)
                if center_rep.assessment_center_branch:
                    queryset = queryset.filter(candidate__assessment_center_branch=center_rep.assessment_center_branch)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def populate_from_candidates(self, request):
        """Populate candidate fees from existing candidates with billing information"""
        created_count = 0
        updated_count = 0
        
        # Get all candidates with assessment series and billing amount
        candidates = Candidate.objects.filter(
            assessment_series__isnull=False
        ).select_related('assessment_series', 'occupation')
        
        for candidate in candidates:
            # Determine total amount based on registration category
            total_amount = 0
            if candidate.registration_category == 'modular' and candidate.modular_billing_amount:
                total_amount = candidate.modular_billing_amount
            elif hasattr(candidate, 'enrollment') and candidate.enrollment.total_amount:
                total_amount = candidate.enrollment.total_amount
            
            # Skip if no billing amount
            if total_amount == 0:
                continue
            
            # Generate payment code
            payment_code = f"{candidate.registration_number}-{candidate.assessment_series.series_code}"
            
            # Determine payment status and amount paid
            amount_paid = candidate.payment_amount_cleared or 0
            payment_status = 'not_paid'
            attempt_status = 'no_attempt'
            payment_date = None
            
            if candidate.payment_cleared_date:
                payment_date = candidate.payment_cleared_date
                if amount_paid >= total_amount:
                    payment_status = 'successful'
                    attempt_status = 'successful'
                elif amount_paid > 0:
                    payment_status = 'pending_approval'
                    attempt_status = 'pending_approval'
            
            # Create or update candidate fee
            fee, created = CandidateFee.objects.update_or_create(
                candidate=candidate,
                assessment_series=candidate.assessment_series,
                defaults={
                    'payment_code': payment_code,
                    'total_amount': total_amount,
                    'amount_paid': amount_paid,
                    'amount_due': total_amount - amount_paid,
                    'payment_date': payment_date,
                    'payment_status': payment_status,
                    'attempt_status': attempt_status,
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return Response({
            'message': f'Successfully populated candidate fees',
            'created': created_count,
            'updated': updated_count,
            'total': created_count + updated_count
        })
    
    @action(detail=False, methods=['post'])
    def mark_as_paid(self, request):
        """Bulk mark selected candidate fees as paid by accounts office"""
        fee_ids = request.data.get('fee_ids', [])
        payment_reference = request.data.get('payment_reference', '')
        
        if not fee_ids:
            return Response({'error': 'No fee IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        if payment_reference not in ('bulk_payment', 'via_schoolpay'):
            return Response({'error': 'Invalid payment reference. Must be bulk_payment or via_schoolpay'}, status=status.HTTP_400_BAD_REQUEST)
        
        fees = CandidateFee.objects.select_related(
            'candidate', 'assessment_series'
        ).filter(id__in=fee_ids, verification_status='pending')
        count = fees.count()
        
        if count == 0:
            return Response({'error': 'No eligible fees found (must be in pending status)'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Only verified candidates can be marked as paid
        unverified = [f.candidate.registration_number for f in fees if f.candidate.verification_status != 'verified']
        if unverified:
            return Response({
                'error': f'Cannot mark as paid: the following candidates are not verified: {", ".join(unverified)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user if request.user.is_authenticated else None
        now = timezone.now()
        
        centers_to_update = set()
        for fee in fees:
            fee.verification_status = 'marked'
            fee.payment_reference = payment_reference
            fee.marked_by = user
            fee.marked_date = now
            fee.amount_paid = fee.total_amount
            fee.payment_date = now
            fee.save()
            if fee.candidate.assessment_center:
                centers_to_update.add((fee.assessment_series_id, fee.candidate.assessment_center_id))
        
        for series_id, center_id in centers_to_update:
            update_center_fee(
                AssessmentSeries.objects.get(pk=series_id),
                AssessmentCenter.objects.get(pk=center_id),
            )
        
        return Response({
            'success': True,
            'message': f'{count} fee(s) marked as paid',
            'count': count,
        })
    
    @action(detail=False, methods=['post'])
    def approve_payment(self, request):
        """Bulk approve payment for fees that have been marked as paid"""
        fee_ids = request.data.get('fee_ids', [])
        
        if not fee_ids:
            return Response({'error': 'No fee IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        fees = CandidateFee.objects.select_related(
            'candidate', 'assessment_series'
        ).filter(id__in=fee_ids, verification_status='marked')
        count = fees.count()
        
        if count == 0:
            return Response({'error': 'No eligible fees found (must be in marked status)'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Only verified candidates can be approved
        unverified = [f.candidate.registration_number for f in fees if f.candidate.verification_status != 'verified']
        if unverified:
            return Response({
                'error': f'Cannot approve payment: the following candidates are not verified: {", ".join(unverified)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user if request.user.is_authenticated else None
        now = timezone.now()
        
        centers_to_update = set()
        for fee in fees:
            fee.verification_status = 'approved'
            fee.approved_by = user
            fee.approved_date = now
            fee.save()
            if fee.candidate.assessment_center:
                centers_to_update.add((fee.assessment_series_id, fee.candidate.assessment_center_id))
        
        for series_id, center_id in centers_to_update:
            update_center_fee(
                AssessmentSeries.objects.get(pk=series_id),
                AssessmentCenter.objects.get(pk=center_id),
            )
        
        return Response({
            'success': True,
            'message': f'{count} fee(s) approved',
            'count': count,
        })


class CenterFeeViewSet(viewsets.ModelViewSet):
    """ViewSet for managing center fees"""
    serializer_class = CenterFeeSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['assessment_series', 'assessment_center']
    search_fields = ['assessment_center__center_name']
    ordering_fields = ['created_at', 'total_amount', 'total_candidates']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Filter center fees by center for center representatives"""
        queryset = CenterFee.objects.select_related(
            'assessment_series', 'assessment_center'
        ).all()
        
        # Filter by center for center representatives
        if self.request.user.is_authenticated and self.request.user.user_type == 'center_representative':
            if hasattr(self.request.user, 'center_rep_profile'):
                center_rep = self.request.user.center_rep_profile
                queryset = queryset.filter(assessment_center=center_rep.assessment_center)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def candidates(self, request, pk=None):
        """Get all candidate fees for a specific center fee (center + series)"""
        center_fee = self.get_object()
        candidate_fees = CandidateFee.objects.filter(
            assessment_series=center_fee.assessment_series,
            candidate__assessment_center=center_fee.assessment_center
        ).select_related('candidate', 'candidate__occupation', 'assessment_series')
        
        serializer = CandidateFeeSerializer(candidate_fees, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def center_invoice(self, request, pk=None):
        """Generate center invoice PDF"""
        center_fee = self.get_object()
        
        # Create response
        invoice_number = f"{center_fee.assessment_center.center_number}-{center_fee.assessment_series.name.replace(' ', '')}-{str(center_fee.id).zfill(3)}"
        filename = f"{center_fee.assessment_center.center_number}_invoice.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        # Create PDF
        buffer = response
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo path
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'uvtab-logo.png')
        
        # Header with logo
        header_data = [
            ['P.O.Box 1499\nEmail: info@uvtab.go.ug', '', 'Tel: +256392002468']
        ]
        
        # Add logo if exists
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1*inch, height=1*inch)
            header_data[0][1] = logo
        
        header_table = Table(header_data, colWidths=[2.5*inch, 2*inch, 2.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph('UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD', title_style))
        elements.append(Paragraph('(UVTAB)', title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Address
        address_style = ParagraphStyle('Address', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)
        elements.append(Paragraph('P.O.Box 1499, Plot 7, Valley Drive, Ntinda-Kyambogo Road', address_style))
        elements.append(Paragraph('Kampala, Uganda | +256392002468 | info@uvtab.go.ug', address_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Invoice number
        invoice_title_style = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.red,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(f'CENTER INVOICE: {invoice_number}', invoice_title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Center details
        detail_style = ParagraphStyle('Detail', parent=styles['Normal'], fontSize=10, spaceAfter=6)
        bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
        
        elements.append(Paragraph('<b>Center Details:</b>', bold_style))
        elements.append(Paragraph(f'Center No.: {center_fee.assessment_center.center_number}', detail_style))
        elements.append(Paragraph(f'Center Name: {center_fee.assessment_center.center_name}', detail_style))
        elements.append(Paragraph(f'Location: {center_fee.assessment_center.get_full_location()}', detail_style))
        elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Paragraph(f'<b>Assessment Series:</b> {center_fee.assessment_series.name}', detail_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Payment details
        elements.append(Paragraph('<b>Payment Details:</b> UVTAB Account <b>9030026294419</b> at <b>Stanbic Bank, Metro Branch</b>.', detail_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Get candidate breakdown by registration category
        candidate_fees = CandidateFee.objects.filter(
            assessment_series=center_fee.assessment_series,
            candidate__assessment_center=center_fee.assessment_center
        ).select_related('candidate')
        
        # Count by registration category
        breakdown = {}
        for fee in candidate_fees:
            reg_cat = fee.candidate.registration_category.capitalize()
            if reg_cat not in breakdown:
                breakdown[reg_cat] = {'count': 0, 'amount': 0}
            breakdown[reg_cat]['count'] += 1
            breakdown[reg_cat]['amount'] += float(fee.total_amount)
        
        # Build table data
        table_data = [
            ['Description', 'Amount (UGX)'],
            ['Number of Candidates', str(center_fee.total_candidates)]
        ]
        
        # Add breakdown by category
        for cat, data in breakdown.items():
            table_data.append([f'{cat} — {data["count"]} candidate(s)', f'{data["amount"]:,.2f}'])
        
        table_data.append(['Total Bill', f'{float(center_fee.total_amount):,.2f}'])
        table_data.append(['Amount Paid', f'{float(center_fee.amount_paid):,.2f}'])
        table_data.append(['Amount Due', f'{float(center_fee.amount_due):,.2f}'])
        
        # Create table
        invoice_table = Table(table_data, colWidths=[4*inch, 2.5*inch])
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(1, 0.95, 0.7)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(invoice_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
        elements.append(Paragraph(f'Generated on {datetime.now().strftime("%B %d, %Y at %I:%M %p")}', footer_style))
        elements.append(Paragraph('UVTAB EIMS - Education Information Management System', footer_style))
        
        # Build PDF
        doc.build(elements)
        return response
    
    @action(detail=True, methods=['get'])
    def candidate_summary_invoice(self, request, pk=None):
        """Generate candidate summary invoice PDF with detailed candidate list"""
        center_fee = self.get_object()
        
        # Create response
        invoice_number = f"{center_fee.assessment_center.center_number}-{center_fee.assessment_series.name.replace(' ', '')}-{str(center_fee.id).zfill(3)}"
        filename = f"{center_fee.assessment_center.center_number}_candidatesummaryinvoice.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        # Create PDF in landscape orientation
        buffer = response
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo path
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'uvtab-logo.png')
        
        # Header with logo
        header_data = [
            ['P.O.Box 1499\nEmail: info@uvtab.go.ug', '', 'Tel: +256392002468']
        ]
        
        # Add logo if exists
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1*inch, height=1*inch)
            header_data[0][1] = logo
        
        header_table = Table(header_data, colWidths=[3.5*inch, 2*inch, 3.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph('UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD (UVTAB)', title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Address
        address_style = ParagraphStyle('Address', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)
        elements.append(Paragraph('P.O.Box 1499, Plot 7, Valley Drive, Ntinda-Kyambogo Road', address_style))
        elements.append(Paragraph('Kampala, Uganda | +256392002468 | info@uvtab.go.ug', address_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Invoice number
        invoice_title_style = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.red,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(f'CENTER INVOICE: {invoice_number}', invoice_title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Center details
        detail_style = ParagraphStyle('Detail', parent=styles['Normal'], fontSize=10, spaceAfter=6)
        bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
        
        elements.append(Paragraph('<b>Center Details:</b>', bold_style))
        elements.append(Paragraph(f'Center No.: {center_fee.assessment_center.center_number}', detail_style))
        elements.append(Paragraph(f'Center Name: {center_fee.assessment_center.center_name}', detail_style))
        elements.append(Paragraph(f'Location: {center_fee.assessment_center.get_full_location()}', detail_style))
        elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Paragraph(f'<b>Assessment Series:</b> {center_fee.assessment_series.name}', detail_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Payment details
        elements.append(Paragraph('<b>Payment Details:</b> UVTAB Account <b>9030026294419</b> at <b>Stanbic Bank, Metro Branch</b>.', detail_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Get candidate breakdown by registration category
        candidate_fees = CandidateFee.objects.filter(
            assessment_series=center_fee.assessment_series,
            candidate__assessment_center=center_fee.assessment_center
        ).select_related('candidate', 'candidate__occupation').order_by('candidate__registration_category', 'candidate__registration_number')
        
        # Count by registration category
        breakdown = {}
        for fee in candidate_fees:
            reg_cat = fee.candidate.registration_category.capitalize()
            if reg_cat not in breakdown:
                breakdown[reg_cat] = {'count': 0, 'amount': 0}
            breakdown[reg_cat]['count'] += 1
            breakdown[reg_cat]['amount'] += float(fee.total_amount)
        
        # Build summary table data
        table_data = [
            ['Description', 'Amount (UGX)'],
            ['Number of Candidates', str(center_fee.total_candidates)]
        ]
        
        # Add breakdown by category
        for cat, data in breakdown.items():
            table_data.append([f'{cat} — {data["count"]} candidate(s)', f'{data["amount"]:,.2f}'])
        
        table_data.append(['Total Bill', f'{float(center_fee.total_amount):,.2f}'])
        table_data.append(['Amount Paid', f'{float(center_fee.amount_paid):,.2f}'])
        table_data.append(['Amount Due', f'{float(center_fee.amount_due):,.2f}'])
        
        # Create summary table (wider for landscape)
        invoice_table = Table(table_data, colWidths=[5*inch, 3*inch])
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(1, 0.95, 0.7)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(invoice_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Add page break to keep summary on first page
        elements.append(PageBreak())
        
        # Candidate Details section
        elements.append(Paragraph('<b>Candidate Details</b>', bold_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Group candidates by registration category
        for cat in sorted(breakdown.keys()):
            elements.append(Paragraph(f'<b>{cat} Candidates Details</b>', bold_style))
            elements.append(Spacer(1, 0.1*inch))
            
            # Filter candidates for this category
            cat_fees = [f for f in candidate_fees if f.candidate.registration_category.capitalize() == cat]
            
            # Build candidate table
            candidate_table_data = [
                ['Reg. Number', 'Name', 'Occupation', 'No. of Modules', 'Billed (UGX)', 'Paid (UGX)', 'Due (UGX)']
            ]
            
            for fee in cat_fees:
                candidate = fee.candidate
                # Get module count from enrollment
                module_count = 0
                try:
                    enrollment = candidate.enrollments.filter(assessment_series=center_fee.assessment_series).first()
                    if enrollment:
                        module_count = enrollment.enrollment_modules.count()
                except:
                    pass
                
                candidate_table_data.append([
                    candidate.registration_number,
                    candidate.full_name,
                    candidate.occupation.occ_name if candidate.occupation else '-',
                    str(module_count) if module_count > 0 else '-',
                    f'{float(fee.total_amount):,.2f}',
                    f'{float(fee.amount_paid):,.2f}',
                    f'{float(fee.amount_due):,.2f}'
                ])
            
            # Create candidate table (adjusted columns for landscape to prevent overlap)
            cand_table = Table(candidate_table_data, colWidths=[1.5*inch, 2.2*inch, 1.8*inch, 0.9*inch, 1.1*inch, 1.1*inch, 1.1*inch])
            cand_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(cand_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Footer
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
        elements.append(Paragraph(f'Generated on {datetime.now().strftime("%B %d, %Y at %I:%M %p")}', footer_style))
        elements.append(Paragraph('UVTAB EIMS - Education Information Management System', footer_style))
        
        # Build PDF
        doc.build(elements)
        return response
    
    @action(detail=False, methods=['get'])
    def quarterly_report(self, request):
        """Generate quarterly financial report as Excel file"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        quarter = request.query_params.get('quarter')
        series_id = request.query_params.get('assessment_series')
        
        if not quarter:
            return Response({'error': 'Quarter is required (Q1, Q2, Q3, Q4)'}, status=status.HTTP_400_BAD_REQUEST)
        if quarter not in ('Q1', 'Q2', 'Q3', 'Q4'):
            return Response({'error': 'Invalid quarter. Must be Q1, Q2, Q3, or Q4'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Build queryset
        center_fees = CenterFee.objects.select_related(
            'assessment_series', 'assessment_center'
        ).filter(assessment_series__quarter=quarter)
        
        if series_id:
            center_fees = center_fees.filter(assessment_series_id=series_id)
        
        if not center_fees.exists():
            return Response({'error': 'No center fee data found for the selected quarter/series'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        quarter_labels = {
            'Q1': 'Q1 (July - September)',
            'Q2': 'Q2 (October - December)',
            'Q3': 'Q3 (January - March)',
            'Q4': 'Q4 (April - June)',
        }
        ws.title = f'{quarter} Report'
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = '#,##0'
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f'UVTAB Quarterly Financial Report — {quarter_labels[quarter]}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Center Code', 'Center Name', 'Total Candidates', 'Assessment Series', 'Amount Billed (UGX)', 'Amount Paid (UGX)', 'Amount Due (UGX)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        row_idx = 4
        total_candidates = 0
        total_billed = 0
        total_paid = 0
        total_due = 0
        
        for cf in center_fees.order_by('assessment_center__center_number', 'assessment_series__name'):
            ws.cell(row=row_idx, column=1, value=cf.assessment_center.center_number).border = border
            ws.cell(row=row_idx, column=2, value=cf.assessment_center.center_name).border = border
            
            cand_cell = ws.cell(row=row_idx, column=3, value=cf.total_candidates)
            cand_cell.border = border
            cand_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_idx, column=4, value=cf.assessment_series.name).border = border
            
            billed_cell = ws.cell(row=row_idx, column=5, value=float(cf.total_amount))
            billed_cell.border = border
            billed_cell.number_format = money_format
            billed_cell.alignment = Alignment(horizontal='right')
            
            paid_cell = ws.cell(row=row_idx, column=6, value=float(cf.amount_paid))
            paid_cell.border = border
            paid_cell.number_format = money_format
            paid_cell.alignment = Alignment(horizontal='right')
            
            due_cell = ws.cell(row=row_idx, column=7, value=float(cf.amount_due))
            due_cell.border = border
            due_cell.number_format = money_format
            due_cell.alignment = Alignment(horizontal='right')
            
            total_candidates += cf.total_candidates
            total_billed += float(cf.total_amount)
            total_paid += float(cf.amount_paid)
            total_due += float(cf.amount_due)
            row_idx += 1
        
        # Totals row
        total_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        total_font = Font(bold=True)
        
        ws.cell(row=row_idx, column=1, value='').border = border
        total_label = ws.cell(row=row_idx, column=2, value='TOTALS')
        total_label.font = total_font
        total_label.border = border
        
        tc = ws.cell(row=row_idx, column=3, value=total_candidates)
        tc.font = total_font
        tc.border = border
        tc.alignment = Alignment(horizontal='center')
        tc.fill = total_fill
        
        ws.cell(row=row_idx, column=4, value='').border = border
        
        tb = ws.cell(row=row_idx, column=5, value=total_billed)
        tb.font = total_font
        tb.border = border
        tb.number_format = money_format
        tb.alignment = Alignment(horizontal='right')
        tb.fill = total_fill
        
        tp = ws.cell(row=row_idx, column=6, value=total_paid)
        tp.font = total_font
        tp.border = border
        tp.number_format = money_format
        tp.alignment = Alignment(horizontal='right')
        tp.fill = total_fill
        
        td = ws.cell(row=row_idx, column=7, value=total_due)
        td.font = total_font
        td.border = border
        td.number_format = money_format
        td.alignment = Alignment(horizontal='right')
        td.fill = total_fill
        
        # Column widths
        col_widths = [15, 40, 18, 28, 20, 20, 20]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Write to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'UVTAB_Quarterly_Report_{quarter}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=False, methods=['post'])
    def populate_from_candidates(self, request):
        """Populate center fees by aggregating candidate fees per center and series"""
        from assessment_centers.models import AssessmentCenter
        from assessment_series.models import AssessmentSeries
        
        created_count = 0
        updated_count = 0
        
        # Get all assessment series
        series_list = AssessmentSeries.objects.all()
        centers = AssessmentCenter.objects.all()
        
        for series in series_list:
            for center in centers:
                # Aggregate candidate fees for this center and series
                candidate_fees = CandidateFee.objects.filter(
                    assessment_series=series,
                    candidate__assessment_center=center
                )
                
                if not candidate_fees.exists():
                    continue
                
                aggregated = candidate_fees.aggregate(
                    total_candidates=Count('id'),
                    total_amount=Sum('total_amount'),
                    amount_paid=Sum('amount_paid')
                )
                
                # Create or update center fee
                fee, created = CenterFee.objects.update_or_create(
                    assessment_series=series,
                    assessment_center=center,
                    defaults={
                        'total_candidates': aggregated['total_candidates'] or 0,
                        'total_amount': aggregated['total_amount'] or 0,
                        'amount_paid': aggregated['amount_paid'] or 0,
                        'amount_due': (aggregated['total_amount'] or 0) - (aggregated['amount_paid'] or 0),
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        return Response({
            'message': f'Successfully populated center fees',
            'created': created_count,
            'updated': updated_count,
            'total': created_count + updated_count
        })
