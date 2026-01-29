"""
Excel Export functionality for assessment series statistics
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from assessment_series.models import AssessmentSeries
from occupations.models import Sector


def create_formatted_excel(series, overview, category_stats, sector_stats, occupation_stats, grade_dist, centers_by_sector, centers_summary):
    """
    Create a formatted Excel workbook with assessment series statistics (Candidate Centric)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Series Results"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:N{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"Assessment Series Results: {series.name}"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    title_cell.alignment = center_alignment
    current_row += 1
    
    # Series Info
    ws[f'A{current_row}'] = "Series Period:"
    ws[f'B{current_row}'] = f"{series.start_date} to {series.end_date}"
    current_row += 2
    
    # Overview Section
    ws.merge_cells(f'A{current_row}:N{current_row}')
    overview_header = ws[f'A{current_row}']
    overview_header.value = "OVERVIEW STATISTICS"
    overview_header.font = header_font
    overview_header.fill = header_fill
    overview_header.alignment = center_alignment
    current_row += 1
    
    # Overview headers
    overview_headers = [
        'Metric', 
        'Enrolled (M)', 'Enrolled (F)', 'Enrolled (Total)',
        'Missing (M)', 'Missing (F)', 'Missing (Total)', 'Missing %',
        'Sat (M)', 'Sat (F)', 'Sat (Total)', 'Sat %',
        'Passed (M)', 'Passed (F)', 'Passed (Total)',
        'Failed (M)', 'Failed (F)', 'Failed (Total)',
        'Pass %', 'Fail %'
    ]
    for col_num, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Overview data
    ws[f'A{current_row}'] = "All Candidates"
    ws[f'B{current_row}'] = overview['male_enrolled']
    ws[f'C{current_row}'] = overview['female_enrolled']
    ws[f'D{current_row}'] = overview['enrolled']
    ws[f'E{current_row}'] = overview['male_missing']
    ws[f'F{current_row}'] = overview['female_missing']
    ws[f'G{current_row}'] = overview['missing']
    ws[f'H{current_row}'] = f"{overview['missing_rate']}%"
    ws[f'I{current_row}'] = overview['male_sat']
    ws[f'J{current_row}'] = overview['female_sat']
    ws[f'K{current_row}'] = overview['sat']
    ws[f'L{current_row}'] = f"{overview['sat_rate']}%"
    ws[f'M{current_row}'] = overview['male_passed']
    ws[f'N{current_row}'] = overview['female_passed']
    ws[f'O{current_row}'] = overview['passed']
    ws[f'P{current_row}'] = overview['male_failed']
    ws[f'Q{current_row}'] = overview['female_failed']
    ws[f'R{current_row}'] = overview['failed']
    ws[f'S{current_row}'] = f"{overview['pass_rate']}%"
    ws[f'T{current_row}'] = f"{overview['fail_rate']}%"
    
    for col in range(1, 21):
        ws.cell(row=current_row, column=col).border = border
    current_row += 3
    
    
    # Centers Registered by Sector
    ws.merge_cells(f'A{current_row}:N{current_row}')
    center_header = ws[f'A{current_row}']
    center_header.value = "CENTERS REGISTERED BY SECTOR"
    center_header.font = header_font
    center_header.fill = header_fill
    center_header.alignment = center_alignment
    current_row += 1
    
    # Centers headers
    ws[f'A{current_row}'] = "Sector"
    ws[f'B{current_row}'] = "Centers"
    ws[f'C{current_row}'] = "Branches"
    
    # Style headers
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Write data
    for sector in centers_by_sector:
        ws[f'A{current_row}'] = sector['name']
        ws[f'B{current_row}'] = sector['centers_count']
        ws[f'C{current_row}'] = sector['branch_count']
        
        # Style row
        for col in range(1, 15): # Apply border to row (A-N)
             ws.cell(row=current_row, column=col).border = border
             
        # Center align count
        ws[f'B{current_row}'].alignment = center_alignment
        ws[f'C{current_row}'].alignment = center_alignment
        
        current_row += 1
        
    # Stats Summary Row
    ws[f'A{current_row}'] = "TOTAL (Unique)"
    ws[f'B{current_row}'] = centers_summary['total_centers']
    ws[f'C{current_row}'] = centers_summary['total_branches']
    
    # Style summary row
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = center_alignment
        cell.border = border
    
    # Border for rest of row
    for col in range(4, 15):
        ws.cell(row=current_row, column=col).border = border
        
    current_row += 2

    # Helper for table generation
    def write_stats_table(title, headers, data_list, name_key='name'):
        nonlocal current_row
        
        ws.merge_cells(f'A{current_row}:T{current_row}')
        cat_header = ws[f'A{current_row}']
        cat_header.value = title
        cat_header.font = header_font
        cat_header.fill = header_fill
        cat_header.alignment = center_alignment
        current_row += 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        summary_font = Font(bold=True)

        for item in data_list:
            ws[f'A{current_row}'] = item[name_key]
            ws[f'B{current_row}'] = item['male_enrolled']
            ws[f'C{current_row}'] = item['female_enrolled']
            ws[f'D{current_row}'] = item['enrolled']
            ws[f'E{current_row}'] = item['male_missing']
            ws[f'F{current_row}'] = item['female_missing']
            ws[f'G{current_row}'] = item['missing']
            ws[f'H{current_row}'] = f"{item['missing_rate']}%"
            ws[f'I{current_row}'] = item['male_sat']
            ws[f'J{current_row}'] = item['female_sat']
            ws[f'K{current_row}'] = item['sat']
            ws[f'L{current_row}'] = f"{item['sat_rate']}%"
            ws[f'M{current_row}'] = item['male_passed']
            ws[f'N{current_row}'] = item['female_passed']
            ws[f'O{current_row}'] = item['passed']
            ws[f'P{current_row}'] = item['male_failed']
            ws[f'Q{current_row}'] = item['female_failed']
            ws[f'R{current_row}'] = item['failed']
            ws[f'S{current_row}'] = f"{item['pass_rate']}%"
            ws[f'T{current_row}'] = f"{item['fail_rate']}%"
            
            is_summary = item.get('name') == 'Total' or item.get('is_sector_summary', False)

            for col in range(1, 21):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if is_summary:
                    cell.fill = summary_fill
                    cell.font = summary_font
            current_row += 1
        current_row += 2

    col_headers = [
        'Category', 
        'Enrolled (M)', 'Enrolled (F)', 'Enrolled (T)',
        'Missing (M)', 'Missing (F)', 'Missing (T)', 'Missing %',
        'Sat (M)', 'Sat (F)', 'Sat (T)', 'Sat %',
        'Passed (M)', 'Passed (F)', 'Passed (T)',
        'Failed (M)', 'Failed (F)', 'Failed (T)',
        'Pass %', 'Fail %'
    ]
    
    # Performance by Category
    # Performance by Category
    # category_stats is now a list with a Total row
    write_stats_table("PERFORMANCE BY REGISTRATION CATEGORY", col_headers, category_stats)
    
    # Performance by Sector
    col_headers[0] = 'Sector'
    write_stats_table("PERFORMANCE BY SECTOR", col_headers, sector_stats, name_key='name')
    
    # Performance by Occupation
    if occupation_stats:
        ws.merge_cells(f'A{current_row}:V{current_row}')
        occ_header = ws[f'A{current_row}']
        occ_header.value = "PERFORMANCE BY OCCUPATION"
        occ_header.font = header_font
        occ_header.fill = header_fill
        occ_header.alignment = center_alignment
        current_row += 1
        
        # Occupation headers
        occ_headers = [
            'Sector', 'Occupation', 'Code', 
            'Enrolled (M)', 'Enrolled (F)', 'Enrolled (T)',
            'Missing (M)', 'Missing (F)', 'Missing (T)', 'Missing %',
            'Sat (M)', 'Sat (F)', 'Sat (T)', 'Sat %',
            'Passed (M)', 'Passed (F)', 'Passed (T)',
            'Failed (M)', 'Failed (F)', 'Failed (T)',
            'Pass %', 'Fail %'
        ]
        for col_num, header in enumerate(occ_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Occupation data
        summary_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        summary_font = Font(bold=True, size=11)
        
        for occ in occupation_stats:
            is_summary = occ.get('is_sector_summary', False)
            
            ws[f'A{current_row}'] = '' if is_summary else occ['sector_name']
            ws[f'B{current_row}'] = occ['occupation_name']
            ws[f'C{current_row}'] = occ['occupation_code']
            ws[f'D{current_row}'] = occ['male_enrolled']
            ws[f'E{current_row}'] = occ['female_enrolled']
            ws[f'F{current_row}'] = occ['enrolled']
            ws[f'G{current_row}'] = occ['male_missing']
            ws[f'H{current_row}'] = occ['female_missing']
            ws[f'I{current_row}'] = occ['missing']
            ws[f'J{current_row}'] = f"{occ['missing_rate']}%"
            ws[f'K{current_row}'] = occ['male_sat']
            ws[f'L{current_row}'] = occ['female_sat']
            ws[f'M{current_row}'] = occ['sat']
            ws[f'N{current_row}'] = f"{occ['sat_rate']}%"
            ws[f'O{current_row}'] = occ['male_passed']
            ws[f'P{current_row}'] = occ['female_passed']
            ws[f'Q{current_row}'] = occ['passed']
            ws[f'R{current_row}'] = occ['male_failed']
            ws[f'S{current_row}'] = occ['female_failed']
            ws[f'T{current_row}'] = occ['failed']
            ws[f'U{current_row}'] = f"{occ['pass_rate']}%"
            ws[f'V{current_row}'] = f"{occ['fail_rate']}%"
            
            # Apply formatting
            for col in range(1, 23):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if is_summary:
                    cell.fill = summary_fill
                    cell.font = summary_font
                
                # Check for GRAND TOTAL
                if occ.get('occupation_name') == 'GRAND TOTAL':
                     cell.fill = summary_fill
                     cell.font = summary_font
            current_row += 1
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

@api_view(['GET'])
@permission_classes([AllowAny])
def export_series_excel(request, series_id):
    """
    Export statistics for a specific assessment series to Excel (Candidate Centric)
    """
    from assessment_series.models import AssessmentSeries
    from .utils import calculate_series_statistics
    
    try:
        # Filter by Centers if provided
        center_ids_param = request.query_params.get('center_ids')
        center_ids = [int(id) for id in center_ids_param.split(',')] if center_ids_param else []
        
        # Calculate stats
        stats_data = calculate_series_statistics(series_id, center_ids)
        
        if not stats_data:
             return Response({'error': 'Series not found'}, status=404)
        
        series_obj = AssessmentSeries.objects.get(id=series_id)
        
        # Create Excel workbook
        wb = create_formatted_excel(
            series=series_obj,
            overview=stats_data['overview'],
            category_stats=stats_data['category_stats'],
            sector_stats=stats_data['sector_stats'],
            occupation_stats=stats_data['occupation_stats'],
            grade_dist=stats_data['grade_distribution'],
            centers_by_sector=stats_data['centers_by_sector'],
            centers_summary=stats_data['centers_by_sector_summary']
        )
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Series_Results_{series_obj.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except AssessmentSeries.DoesNotExist:
        return Response({'error': 'Series not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)

def create_special_needs_excel(special_needs_overview, disability_breakdown, special_needs_sector, 
                                refugee_overview, refugee_sector, series_filter=None):
    """
    Create a formatted Excel workbook for special needs statistics
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Special Needs Analytics"
    
    # Define styles
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:K{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "Special Needs Candidates Analytics"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    title_cell.alignment = center_alignment
    current_row += 1
    
    # Filter info
    if series_filter:
        ws[f'A{current_row}'] = f"Filtered by: {series_filter}"
        current_row += 1
    else:
        ws[f'A{current_row}'] = "Showing: All Assessment Series"
        current_row += 1
    current_row += 1
    
    # Overview Section
    ws.merge_cells(f'A{current_row}:K{current_row}')
    overview_header = ws[f'A{current_row}']
    overview_header.value = "OVERALL STATISTICS"
    overview_header.font = header_font
    overview_header.fill = header_fill
    overview_header.alignment = center_alignment
    current_row += 1
    
    # Overview headers
    overview_headers = ['Metric', 'Total', 'Male', 'Female', 'Male Passed', 'Female Passed', 
                        'Total Passed', 'Male Pass %', 'Female Pass %', 'Overall Pass %']
    for col_num, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Overview data
    ws[f'A{current_row}'] = "Special Needs Candidates"
    ws[f'B{current_row}'] = special_needs_overview['total']
    ws[f'C{current_row}'] = special_needs_overview['male']
    ws[f'D{current_row}'] = special_needs_overview['female']
    ws[f'E{current_row}'] = special_needs_overview['male_passed']
    ws[f'F{current_row}'] = special_needs_overview['female_passed']
    ws[f'G{current_row}'] = special_needs_overview['total_passed']
    ws[f'H{current_row}'] = f"{special_needs_overview['male_pass_rate']}%"
    ws[f'I{current_row}'] = f"{special_needs_overview['female_pass_rate']}%"
    ws[f'J{current_row}'] = f"{special_needs_overview['pass_rate']}%"
    
    for col in range(1, 11):
        ws.cell(row=current_row, column=col).border = border
    current_row += 3
    
    # Performance by Disability Type
    if disability_breakdown:
        ws.merge_cells(f'A{current_row}:K{current_row}')
        dis_header = ws[f'A{current_row}']
        dis_header.value = "PERFORMANCE BY DISABILITY TYPE"
        dis_header.font = header_font
        dis_header.fill = header_fill
        dis_header.alignment = center_alignment
        current_row += 1
        
        # Disability headers
        dis_headers = ['Disability Type', 'Total', 'Male', 'Female', 'Male Passed', 'Female Passed',
                       'Total Passed', 'Male Pass %', 'Female Pass %', 'Overall Pass %']
        for col_num, header in enumerate(dis_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Disability data
        for disability in disability_breakdown:
            ws[f'A{current_row}'] = disability['name']
            ws[f'B{current_row}'] = disability['total']
            ws[f'C{current_row}'] = disability['male']
            ws[f'D{current_row}'] = disability['female']
            ws[f'E{current_row}'] = disability['male_passed']
            ws[f'F{current_row}'] = disability['female_passed']
            ws[f'G{current_row}'] = disability['total_passed']
            ws[f'H{current_row}'] = f"{disability['male_pass_rate']}%"
            ws[f'I{current_row}'] = f"{disability['female_pass_rate']}%"
            ws[f'J{current_row}'] = f"{disability['pass_rate']}%"
            
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        current_row += 2
    
    # Performance by Sector - Special Needs
    if special_needs_sector:
        ws.merge_cells(f'A{current_row}:K{current_row}')
        sector_header = ws[f'A{current_row}']
        sector_header.value = "PERFORMANCE BY SECTOR"
        sector_header.font = header_font
        sector_header.fill = header_fill
        sector_header.alignment = center_alignment
        current_row += 1
        
        # Sector headers
        sector_headers = ['Sector Name', 'Total', 'Male', 'Female', 'Male Passed', 'Female Passed',
                         'Total Passed', 'Male Pass %', 'Female Pass %', 'Overall Pass %']
        for col_num, header in enumerate(sector_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Sector data
        for sector in special_needs_sector:
            ws[f'A{current_row}'] = sector['sector_name']
            ws[f'B{current_row}'] = sector['total']
            ws[f'C{current_row}'] = sector['male']
            ws[f'D{current_row}'] = sector['female']
            ws[f'E{current_row}'] = sector['male_passed']
            ws[f'F{current_row}'] = sector['female_passed']
            ws[f'G{current_row}'] = sector['total_passed']
            ws[f'H{current_row}'] = f"{sector['male_pass_rate']}%"
            ws[f'I{current_row}'] = f"{sector['female_pass_rate']}%"
            ws[f'J{current_row}'] = f"{sector['pass_rate']}%"
            
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
    
    # REFUGEE STATISTICS SECTION
    current_row += 2
    
    # Refugee Overview
    ws.merge_cells(f'A{current_row}:K{current_row}')
    refugee_header = ws[f'A{current_row}']
    refugee_header.value = "REFUGEE CANDIDATES"
    refugee_header.font = header_font
    refugee_header.fill = header_fill
    refugee_header.alignment = center_alignment
    current_row += 1
    
    # Refugee overview headers (reuse same format)
    for col_num, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    current_row += 1
    
    # Refugee overview data
    ws[f'A{current_row}'] = "Refugee Candidates"
    ws[f'B{current_row}'] = refugee_overview['total']
    ws[f'C{current_row}'] = refugee_overview['male']
    ws[f'D{current_row}'] = refugee_overview['female']
    ws[f'E{current_row}'] = refugee_overview['male_passed']
    ws[f'F{current_row}'] = refugee_overview['female_passed']
    ws[f'G{current_row}'] = refugee_overview['total_passed']
    ws[f'H{current_row}'] = f"{refugee_overview['male_pass_rate']}%"
    ws[f'I{current_row}'] = f"{refugee_overview['female_pass_rate']}%"
    ws[f'J{current_row}'] = f"{refugee_overview['pass_rate']}%"
    
    for col in range(1, 11):
        ws.cell(row=current_row, column=col).border = border
    current_row += 3
    
    # Performance by Sector - Refugee
    if refugee_sector:
        ws.merge_cells(f'A{current_row}:K{current_row}')
        sector_header = ws[f'A{current_row}']
        sector_header.value = "REFUGEE PERFORMANCE BY SECTOR"
        sector_header.font = header_font
        sector_header.fill = header_fill
        sector_header.alignment = center_alignment
        current_row += 1
        
        # Sector headers
        sector_headers = ['Sector Name', 'Total', 'Male', 'Female', 'Male Passed', 'Female Passed',
                         'Total Passed', 'Male Pass %', 'Female Pass %', 'Overall Pass %']
        for col_num, header in enumerate(sector_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # Refugee sector data
        for sector in refugee_sector:
            ws[f'A{current_row}'] = sector['sector_name']
            ws[f'B{current_row}'] = sector['total']
            ws[f'C{current_row}'] = sector['male']
            ws[f'D{current_row}'] = sector['female']
            ws[f'E{current_row}'] = sector['male_passed']
            ws[f'F{current_row}'] = sector['female_passed']
            ws[f'G{current_row}'] = sector['total_passed']
            ws[f'H{current_row}'] = f"{sector['male_pass_rate']}%"
            ws[f'I{current_row}'] = f"{sector['female_pass_rate']}%"
            ws[f'J{current_row}'] = f"{sector['pass_rate']}%"
            
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


@api_view(['GET'])
@permission_classes([AllowAny])
def export_special_needs_excel(request):
    """
    Export special needs statistics to Excel file
    """
    try:
        # We need to recreate the logic from special_needs_analytics to avoid request type issues
        from results.models import ModularResult, FormalResult, WorkersPasResult
        from configurations.models import NatureOfDisability
        
        series_id = request.query_params.get('series_id')
        
        # Get all results (optionally filtered by series)
        if series_id:
            all_results = list(ModularResult.objects.filter(assessment_series_id=series_id).select_related('candidate', 'candidate__nature_of_disability')) + \
                          list(FormalResult.objects.filter(assessment_series_id=series_id).select_related('candidate', 'candidate__nature_of_disability')) + \
                          list(WorkersPasResult.objects.filter(assessment_series_id=series_id).select_related('candidate', 'candidate__nature_of_disability'))
            series = AssessmentSeries.objects.get(id=series_id)
            series_filter = series.name
        else:
            all_results = list(ModularResult.objects.all().select_related('candidate', 'candidate__nature_of_disability')) + \
                          list(FormalResult.objects.all().select_related('candidate', 'candidate__nature_of_disability')) + \
                          list(WorkersPasResult.objects.all().select_related('candidate', 'candidate__nature_of_disability'))
            series_filter = None
        
        # Filter to only special needs candidates
        special_needs_results = [r for r in all_results if r.candidate.has_disability]
        
        # Calculate overview
        special_needs_male = [r for r in special_needs_results if r.candidate.gender == 'male']
        special_needs_female = [r for r in special_needs_results if r.candidate.gender == 'female']
        
        def is_passed(result):
            # Skip results without marks
            if result.mark is None:
                return False
                
            if isinstance(result, ModularResult):
                return result.mark >= 65
            elif isinstance(result, FormalResult):
                return (result.type == 'theory' and result.mark >= 50) or (result.type == 'practical' and result.mark >= 65)
            elif isinstance(result, WorkersPasResult):
                return result.mark >= 65
            return False
        
        male_passed = sum(1 for r in special_needs_male if is_passed(r))
        female_passed = sum(1 for r in special_needs_female if is_passed(r))
        total_passed = male_passed + female_passed
        
        overview = {
            'total': len(special_needs_results),
            'male': len(special_needs_male),
            'female': len(special_needs_female),
            'male_passed': male_passed,
            'female_passed': female_passed,
            'total_passed': total_passed,
            'male_pass_rate': round((male_passed / len(special_needs_male) * 100), 2) if len(special_needs_male) > 0 else 0,
            'female_pass_rate': round((female_passed / len(special_needs_female) * 100), 2) if len(special_needs_female) > 0 else 0,
            'pass_rate': round((total_passed / len(special_needs_results) * 100), 2) if len(special_needs_results) > 0 else 0
        }
        
        # By disability type
        disability_breakdown = []
        for disability in NatureOfDisability.objects.all():
            disability_results = [r for r in special_needs_results if r.candidate.nature_of_disability == disability]
            
            if not disability_results:
                continue
            
            dis_male = [r for r in disability_results if r.candidate.gender == 'male']
            dis_female = [r for r in disability_results if r.candidate.gender == 'female']
            
            dis_male_passed = sum(1 for r in dis_male if is_passed(r))
            dis_female_passed = sum(1 for r in dis_female if is_passed(r))
            dis_total_passed = dis_male_passed + dis_female_passed
            
            disability_breakdown.append({
                'name': disability.name,
                'total': len(disability_results),
                'male': len(dis_male),
                'female': len(dis_female),
                'male_passed': dis_male_passed,
                'female_passed': dis_female_passed,
                'total_passed': dis_total_passed,
                'male_pass_rate': round((dis_male_passed / len(dis_male) * 100), 2) if len(dis_male) > 0 else 0,
                'female_pass_rate': round((dis_female_passed / len(dis_female) * 100), 2) if len(dis_female) > 0 else 0,
                'pass_rate': round((dis_total_passed / len(disability_results) * 100), 2) if len(disability_results) > 0 else 0
            })
        
        # By sector
        from occupations.models import Sector
        sector_breakdown = []
        for sector in Sector.objects.all():
            sector_results = [r for r in special_needs_results if 
                             r.candidate.occupation and r.candidate.occupation.sector == sector]
            
            if not sector_results:
                continue
            
            sector_male = [r for r in sector_results if r.candidate.gender == 'male']
            sector_female = [r for r in sector_results if r.candidate.gender == 'female']
            
            sector_male_passed = sum(1 for r in sector_male if is_passed(r))
            sector_female_passed = sum(1 for r in sector_female if is_passed(r))
            sector_total_passed = sector_male_passed + sector_female_passed
            
            sector_breakdown.append({
                'sector_name': sector.name,
                'total': len(sector_results),
                'male': len(sector_male),
                'female': len(sector_female),
                'male_passed': sector_male_passed,
                'female_passed': sector_female_passed,
                'total_passed': sector_total_passed,
                'male_pass_rate': round((sector_male_passed / len(sector_male) * 100), 2) if len(sector_male) > 0 else 0,
                'female_pass_rate': round((sector_female_passed / len(sector_female) * 100), 2) if len(sector_female) > 0 else 0,
                'pass_rate': round((sector_total_passed / len(sector_results) * 100), 2) if len(sector_results) > 0 else 0
            })
        
        # REFUGEE STATISTICS
        refugee_results = [r for r in all_results if r.candidate.is_refugee]
        refugee_male = [r for r in refugee_results if r.candidate.gender == 'male']
        refugee_female = [r for r in refugee_results if r.candidate.gender == 'female']
        
        refugee_male_passed = sum(1 for r in refugee_male if is_passed(r))
        refugee_female_passed = sum(1 for r in refugee_female if is_passed(r))
        refugee_total_passed = refugee_male_passed + refugee_female_passed
        
        refugee_overview = {
            'total': len(refugee_results),
            'male': len(refugee_male),
            'female': len(refugee_female),
            'male_passed': refugee_male_passed,
            'female_passed': refugee_female_passed,
            'total_passed': refugee_total_passed,
            'male_pass_rate': round((refugee_male_passed / len(refugee_male) * 100), 2) if len(refugee_male) > 0 else 0,
            'female_pass_rate': round((refugee_female_passed / len(refugee_female) * 100), 2) if len(refugee_female) > 0 else 0,
            'pass_rate': round((refugee_total_passed / len(refugee_results) * 100), 2) if len(refugee_results) > 0 else 0
        }
        
        # Refugee by sector
        refugee_sector = []
        for sector in Sector.objects.all():
            sector_results = [r for r in refugee_results if 
                             r.candidate.occupation and r.candidate.occupation.sector == sector]
            
            if not sector_results:
                continue
            
            sector_male = [r for r in sector_results if r.candidate.gender == 'male']
            sector_female = [r for r in sector_results if r.candidate.gender == 'female']
            
            sector_male_passed = sum(1 for r in sector_male if is_passed(r))
            sector_female_passed = sum(1 for r in sector_female if is_passed(r))
            sector_total_passed = sector_male_passed + sector_female_passed
            
            refugee_sector.append({
                'sector_name': sector.name,
                'total': len(sector_results),
                'male': len(sector_male),
                'female': len(sector_female),
                'male_passed': sector_male_passed,
                'female_passed': sector_female_passed,
                'total_passed': sector_total_passed,
                'male_pass_rate': round((sector_male_passed / len(sector_male) * 100), 2) if len(sector_male) > 0 else 0,
                'female_pass_rate': round((sector_female_passed / len(sector_female) * 100), 2) if len(sector_female) > 0 else 0,
                'pass_rate': round((sector_total_passed / len(sector_results) * 100), 2) if len(sector_results) > 0 else 0
            })
        
        # Create Excel workbook with both special needs and refugee data
        wb = create_special_needs_excel(
            special_needs_overview=overview,
            disability_breakdown=disability_breakdown,
            special_needs_sector=sector_breakdown,
            refugee_overview=refugee_overview,
            refugee_sector=refugee_sector,
            series_filter=series_filter
        )
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Special_Needs_Analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except AssessmentSeries.DoesNotExist:
        return Response({'error': 'Series not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)
